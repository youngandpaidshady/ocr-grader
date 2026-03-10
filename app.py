# -*- coding: utf-8 -*-
import os
import base64
import json
import logging
from flask import Flask, request, jsonify, render_template, send_file, Response, stream_with_context, make_response
from flask_cors import CORS
import google.generativeai as genai
import pandas as pd
from thefuzz import process, fuzz
from dotenv import load_dotenv
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func

# Load environment variables
import time
import re as re_mod
import glob
from concurrent.futures import ThreadPoolExecutor, as_completed
load_dotenv()

# ═══════════════════════════════════════════════════════════════
#  PRODUCTION CONSTANTS
# ═══════════════════════════════════════════════════════════════
AI_MODEL_PRIMARY = 'gemini-2.5-flash'
AI_MODEL_FALLBACK = 'gemini-2.0-flash'
AI_MAX_RETRIES = 3
AI_BACKOFF_SECONDS = [3, 8, 15]  # Exponential backoff per attempt
FUZZY_MATCH_THRESHOLD = 95  # Minimum score for class name fuzzy matching
MAX_UPLOAD_MB = 100
MAX_CONVERSATION_HISTORY = 20  # Sliding window for chat context
MAX_STUDENTS_IN_CONTEXT = 20  # Students sent to AI prompt per class

# ═══════════════════════════════════════════════════════════════
#  STRUCTURED LOGGING
# ═══════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('smartgrader')

# Active working file path configured globally
WORKING_EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ActiveRoaster.xlsx")

# Configure AI Model — supports multiple API keys (comma-separated) for rotation
import threading
_api_keys_raw = os.getenv("GEMINI_API_KEY", "")
API_KEYS = [k.strip() for k in _api_keys_raw.split(",") if k.strip() and k.strip() != "your_gemini_api_key_here"]
_current_key_index = 0
_key_lock = threading.Lock()

if not API_KEYS:
    logger.warning("No API keys set in .env file. Set GEMINI_API_KEY (comma-separated for multiple).")
else:
    logger.info("Loaded {} API key(s) for rotation.".format(len(API_KEYS)))

def get_current_api_key():
    """Get the current API key."""
    if not API_KEYS:
        return None
    return API_KEYS[_current_key_index % len(API_KEYS)]

def rotate_api_key():
    """Rotate to the next API key. Thread-safe."""
    global _current_key_index
    if len(API_KEYS) <= 1:
        return get_current_api_key()
    with _key_lock:
        _current_key_index = (_current_key_index + 1) % len(API_KEYS)
        new_key = API_KEYS[_current_key_index]
        genai.configure(api_key=new_key)
        logger.info("Rotated to API key #{} of {}".format(_current_key_index + 1, len(API_KEYS)))
    return new_key

def _call_gemini(model_name, content_parts, max_retries=None):
    """Centralized Gemini API caller with retry, rotation, and timeout.
    Returns raw_text on success, raises on total failure."""
    retries = max_retries or AI_MAX_RETRIES
    last_error = None
    for attempt in range(retries):
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(content_parts)
            # Extract text — handle thinking mode responses (skip thought blocks)
            raw_text = ''
            if hasattr(response, 'candidates') and response.candidates:
                for part in response.candidates[0].content.parts:
                    if hasattr(part, 'thought') and part.thought:
                        continue
                    if hasattr(part, 'text') and part.text:
                        raw_text += part.text
            if not raw_text:
                raw_text = response.text.strip()
            raw_text = raw_text.strip()
            if raw_text:
                return raw_text
        except Exception as err:
            last_error = err
            err_str = str(err).lower()
            logger.warning("Gemini {} attempt {}/{} failed: {}".format(model_name, attempt + 1, retries, err))
            if 'quota' in err_str or 'rate' in err_str or '429' in err_str or 'resource' in err_str:
                rotate_api_key()
                if attempt < retries - 1:
                    wait = AI_BACKOFF_SECONDS[min(attempt, len(AI_BACKOFF_SECONDS) - 1)]
                    logger.info("Rate limited — waiting {}s before retry...".format(wait))
                    time.sleep(wait)
            else:
                break  # Non-rate-limit error, don't retry
    raise last_error or Exception("All AI attempts failed")

# Configure with the first key
genai.configure(api_key=get_current_api_key())

# Initialize Flask App
app = Flask(__name__)

# CORS: allow the Render domain + localhost for dev
_render_url = os.getenv('RENDER_EXTERNAL_URL', '')
_allowed_origins = ['http://localhost:5000', 'http://127.0.0.1:5000']
if _render_url:
    _allowed_origins.append(_render_url)
CORS(app, origins=_allowed_origins)

app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

# Configure Database
db_url = os.getenv("DATABASE_URL")
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

if not db_url:
    # On Render, we mount a persistent disk to /data. Locally, we just use the root folder.
    if os.path.exists('/data'):
        db_url = 'sqlite:////data/smartgrader.db'
    else:
        db_url = 'sqlite:///smartgrader.db'

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Database Models
class ClassModel(db.Model):
    __tablename__ = 'classes'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    students = db.relationship('StudentModel', backref='class_obj', lazy=True, cascade="all, delete-orphan")

class StudentModel(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('classes.id'), nullable=False)
    scores = db.relationship('ScoreModel', backref='student_obj', lazy=True, cascade="all, delete-orphan")
    enrollments = db.relationship('EnrollmentModel', backref='student_obj', lazy=True, cascade="all, delete-orphan")

class EnrollmentModel(db.Model):
    __tablename__ = 'enrollments'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    subject_name = db.Column(db.String(100), nullable=False)

class ScoreModel(db.Model):
    __tablename__ = 'scores'
    id = db.Column(db.Integer, primary_key=True)
    score_value = db.Column(db.String(50), nullable=False)
    assessment_type = db.Column(db.String(100), default='Score')
    subject_name = db.Column(db.String(100), default='Uncategorized', server_default='Uncategorized')
    term = db.Column(db.String(20), default='1st Term', server_default='1st Term')
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)

from sqlalchemy import text
with app.app_context():
    db.create_all()
    # Migrate: add subject_name column if missing
    try:
        db.session.execute(text("ALTER TABLE scores ADD COLUMN subject_name VARCHAR(100) DEFAULT 'Uncategorized'"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    # Migrate: add term column if missing
    try:
        db.session.execute(text("ALTER TABLE scores ADD COLUMN term VARCHAR(20) DEFAULT '1st Term'"))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # --- ONE-TIME RENDER ROSTER SYNC ---
    # Automatically syncs the DB with the definitive rosters on app startup.
    # This removes phantom students (from old OCR bugs) and adds any missing ones.
    DEFINITIVE_ROSTERS = {
        "SS 1Q": ["Abdul Azeez Haliyah", "Abdul Hammed Qowiyat", "Abdul Kareem Zainab", "Abdulmumin Soburat", "Abdul Quadri Faridah", "Adeaga Mercy", "Adebayo Faouziyah", "Adekola Ayomide", "Adekunle Favour", "Adesina Adeola", "Ajao Lois", "Ajayi Elizabeth", "Ajayi Temilade", "Ajetunmobi Hepzibah", "Akindele Faith", "Akinlekan Faith", "Akinlolu Oluwademilade", "Akpan Priscilla", "Alade Pelumi", "Alaka Courage", "Alli Oluwadarasimi", "Asalu Rukayat", "Asiwaju Alimat", "Atolagbe Deborah", "Bamgbose Faridah", "Bamigbola Fiyinfoluwa", "Bashiru Aishat", "Bello Azeezat", "Benjamin Victoria", "Daniel Oluwadarasimi", "Durodola Naimot", "Edegbai Joy", "Fafowora Mofeyisara", "Fayemi Sunmisola", "Hammed Uruwat", "Hamzat Darasimi", "Ibrahim Yussiroh", "Ilori Julianah", "Jimoh Davida", "Jose Olayinka", "Joshua Tamilore", "Kazeem Alimot", "Kolawole Pemisire", "Lawal Aramide", "Lukman Ayomikun", "Morakinyo Eniola", "Moshood Zainab", "Nurudeen Moridyyah", "Obasuyi Peculiar", "Odesola Suudat", "Okanlawon Esther", "Okeowo Sofiat", "Oladimeji Adijat", "Oladokun Haliyah", "Olaniyi Latifat", "Olanrewaju Aishat", "Olateju Iyanuoluwa", "Olatunde Abigeal", "Olayanju Oluwasikemi", "Olayiwola Fathia", "Oluokun Favor", "Oluwagbenga Deborah", "Oluwasola Oluwaseyifunmi", "Oluwatoyin Jewel", "Oluyemi Oluwadarasimi", "Oriade Bolade", "Salami Oluwatobi", "Unorbueze Victoria", "Uzogbo Chioma"],
        "SS 1S": ["Abdul Kareem Hikmat", "Abdullahi Istijabah", "Abdullateef Olaitan", "Abdulsalam Aishat", "Abdulsalam Sofiat", "Adebayo Inioluwa", "Adefisoye Mariam", "Adejumo Rayhanat", "Adeleke Ifeoluwa", "Adeleke Rebecca", "Adeniji Olamide", "Adepoju Victoria", "Adeshina Adedoyinsola", "Adewole Precious", "Adunse Ibunkunoluwa", "Ahmad Aishah", "Aina Omolewq", "Ajayi Itunuoluwa", "Ajayi Testimony", "Ajibola Olatunmininu", "Akimbo Mercy", "Akindele Habibat", "Alamuoye Seun", "Aminullahi Hikmat", "Amoo Adesewa", "Amuda Khadijah", "Arotile Fortune", "Attai Favour", "Awe Anjola", "Ayangbade Mojirayo", "Ayeloja Yesiroh", "Ayeni Victoria", "Babarinde Adunola", "Babasessin Boluwatife", "Bakare Darasimi", "Ebenezer Marvelous", "Elesho Deborah", "Fakayode Gbemisola", "Fatade Moyinoluwa", "Hoseni Hazeenat", "Ibitoye Oluwadamilola", "Ike Happiness", "Jegede Mercy", "Joel Tofunmi", "Josiah Favour", "Kazeem Faheedah", "Komolafe Ibukun", "Lawal Mazeedah", "Mogaji Asian", "Ocheje Anita", "Odedele Blessing", "Odunbanjo Fariah", "Ogunleye Saidah", "Ogunseye Precious", "Okunade Mercy", "Oladipupo Alimat", "Olapade Victoria", "Olagbegi Ewaoluwa", "Olayiwola Haliyah", "Olowookere Faridah", "Oluborode Semilore", "Oluwadamilare Precious", "Oluwatosin Dorcas", "Oriade Bolanle", "Osineye Mary", "Oyedotun Habibat", "Safiriyu Faizah", "Sakirullahi Aishah", "Sulaimon Haliyah", "Tawhid Mufliah", "Wale Gbadamosi Aisha", "Yakubu Qowiyah"],
        "SS 1I": ["Abass Aishat", "Abdulazeez Zainab", "Adamson Zainab", "Adebayo Ayomide", "Adedeji Lydia", "Adegoke Fiyinfoluwa", "Adegun Aliyat", "Ademola Aishat", "Adeshina Inioluwa", "Adewale Aliyat", "Adeyanju Edith", "Adeyemo Princess", "Agbaje Zainab", "Aina Moridiyah", "Ajani Folashade", "Ajayi Abigal", "Ajiboye Adebimpe", "Ajike Esther", "Ajise Joyce", "Akerele Ewaoluwa", "Alamoye Seyi", "Aluko Irebami", "Amuda Fatimoh", "Apanpa Rokeebat", "Ashabi Ifeoluwa", "Balogun Noimot", "Bello Fatimoh", "Bello Tolulope", "Chukwu Amarachi", "Dada Rukayah", "Daniel Joy", "Daud Hakimat", "Ehidiamin Esther", "Enyi Glorious", "Fajuyigbe Mercy", "Griffin Mushinat", "Hassan Kareemah", "Ibrahim Azeezat", "Ilesanmi Deborah", "Ismail Mariam", "Jacob Favour", "Jegede Mojolaoluwa", "Joseph Adeola", "Kolawole Victoria", "Lawal Kehinde", "Manasseh Peace", "Muideen Sofiat", "Nana Hannah", "Nwaocha Ngozi", "Obideyi Opeoluwa", "Ogunbowale Rokeebat", "Ojo Elizabeth", "Ojo Nifemi", "Okeronbi Oluwadarasimi", "Okoronkwo Grace", "Okunola Arinola", "Oladejo Oluwafeyikemi", "Oladele Inioluwa", "Oladokun Naimot", "Olaitan Oluwadarasimi", "Olaleye Esther", "Olaniyi Abigeal", "Olusola Jesunifemi", "Olutoki Rokeebat", "Oyedeji Olorunwa", "Popoola Esther", "Rafiu Mazeedat", "Rasheed Mosope", "Smith Celestina", "Sulaimon Mariam", "Taiwo Erijuwon"]
    }
    
    try:
        for class_name, correct_names in DEFINITIVE_ROSTERS.items():
            # Create class if missing
            c = ClassModel.query.filter(func.lower(ClassModel.name) == class_name.lower()).first()
            if not c:
                c = ClassModel(name=class_name)
                db.session.add(c)
                db.session.flush()
            
            # Get existing students
            existing = StudentModel.query.filter_by(class_id=c.id).all()
            existing_map = {s.name.strip().lower(): s for s in existing}
            
            # Target list
            target_map = {' '.join(n.strip().split()).lower(): ' '.join(n.strip().split()).title() for n in correct_names}
            
            # 1. Remove extras (like the 3 phantom students in SS 1Q)
            for s in existing:
                s_lower = ' '.join(s.name.strip().split()).lower()
                if s_lower not in target_map:
                    logger.info("Startup Sync [{}]: Removed phantom student '{}'".format(class_name, s.name))
                    db.session.delete(s)
                elif s.name != target_map[s_lower]:
                    # Update capitalization / spacing to be perfectly uniform
                    s.name = target_map[s_lower]
                    
            # 2. Add missing
            for t_lower, t_title in target_map.items():
                if t_lower not in existing_map:
                    logger.info("Startup Sync [{}]: Added missing student '{}'".format(class_name, t_title))
                    db.session.add(StudentModel(name=t_title, class_id=c.id))
                    
        db.session.commit()
    except Exception as e:
        logger.error("Startup Roster Sync Error: {}".format(e))
        db.session.rollback()

# ═══════════════════════════════════════════════════════════════
#  NIGERIAN MARK BOOK GRADING ENGINE
#  Central config for Nigerian school record sheet structure
# ═══════════════════════════════════════════════════════════════
NIGERIAN_MARK_BOOK_CONFIG = {
    "ca_columns": {
        "1st CA":      {"max": 10, "can_be_20": False, "aliases": ["1st Test", "1st C.A", "CA1", "Test 1", "First CA"]},
        "2nd CA":      {"max": 10, "can_be_20": False, "aliases": ["2nd Test", "2nd C.A", "CA2", "Test 2", "Second CA"]},
        "Open Day":    {"max": 20, "can_be_20": True,  "aliases": ["Open", "OD", "Open day"]},
        "Note Book":   {"max": 10, "can_be_20": True,  "aliases": ["Note", "NB", "Notebook", "Note book"]},
        "Assignment":  {"max": 10, "can_be_20": True,  "aliases": ["Ass", "Assig", "Assgn", "Assign"]},
    },
    "ca_total_max": 30,       # Sum of CAs ÷ 2
    "ca_raw_max": 60,         # Sum before dividing
    "exam_max": 70,
    "grand_total_max": 100,
    "absent_markers": ["AB", "ABS", "A", "-", "/", "NIL", "", "X", "ABSENT"],
    "grade_map": [
        (75, 100, "A1", "Excellent"),
        (70, 74,  "B2", "Very Good"),
        (65, 69,  "B3", "Good"),
        (60, 64,  "C4", "Credit"),
        (55, 59,  "C5", "Credit"),
        (50, 54,  "C6", "Credit"),
        (45, 49,  "D7", "Pass"),
        (40, 44,  "E8", "Pass"),
        (0,  39,  "F9", "Fail"),
    ],
    "assessment_order": [
        "1st CA", "2nd CA", "Open Day", "Note Book",
        "Assignment", "Attendance", "Total CA", "Exam", "Grand Total"
    ],
    "column_3_default": "Open Day",
    "terms": ["1st Term", "2nd Term", "3rd Term"],
}

def normalize_column_name(col_name, config=None):
    """Normalize a column name to its canonical form using aliases."""
    if config is None:
        config = NIGERIAN_MARK_BOOK_CONFIG
    col_upper = str(col_name).strip().upper()
    for canonical, info in config["ca_columns"].items():
        if col_upper == canonical.upper():
            return canonical
        for alias in info.get("aliases", []):
            if col_upper == alias.upper():
                return canonical
    # Check Exam
    if col_upper in ["EXAM", "EXAMINATION", "FINAL EXAM", "EXAM SCORE"]:
        return "Exam"
    # Check totals
    if col_upper in ["TOTAL CA", "CA TOTAL", "TOTAL C.A", "TOTAL"]:
        return "Total CA"
    if col_upper in ["GRAND TOTAL", "TOTAL SCORE", "FINAL TOTAL", "G.TOTAL", "OVERALL"]:
        return "Grand Total"
    return col_name.strip()

def validate_and_cap_score(column_name, value, config=None):
    """Validate a score against its column's maximum.
    Returns (cleaned_value, warnings_list).
    - Absent markers → ('ABS', [])
    - Fractions (½) → converted to .5
    - x/y format → numerator extracted
    - Over-max → capped with warning
    """
    if config is None:
        config = NIGERIAN_MARK_BOOK_CONFIG
    warnings = []
    val_str = str(value).strip()

    # Check absent markers
    if val_str.upper() in [m.upper() for m in config["absent_markers"]]:
        return ("ABS", [])

    # Handle fractions: 6½ → 6.5, 8½ → 8.5
    val_str = val_str.replace('½', '.5').replace('¼', '.25').replace('¾', '.75')

    # Handle x/y format: "8/10" → "8"
    if '/' in val_str:
        parts = val_str.split('/')
        val_str = parts[0].strip()

    # Try to parse as number
    try:
        numeric_val = float(val_str)
    except (ValueError, TypeError):
        return (val_str, ["Could not parse '{}' as a number for {}".format(value, column_name)])

    # Determine the max for this column
    col_normalized = normalize_column_name(column_name, config)
    max_val = None

    if col_normalized in config["ca_columns"]:
        ca_info = config["ca_columns"][col_normalized]
        # Use 20 if can_be_20 is True AND the value suggests it (>10)
        if ca_info["can_be_20"] and numeric_val > 10:
            max_val = 20
        else:
            max_val = ca_info["max"]
    elif col_normalized == "Total CA":
        max_val = config["ca_total_max"]
    elif col_normalized == "Exam":
        max_val = config["exam_max"]
    elif col_normalized in ["Grand Total", "Total Score"]:
        max_val = config["grand_total_max"]

    if max_val is not None and numeric_val > max_val:
        warnings.append("{} score {} exceeds max {} — capped".format(column_name, numeric_val, max_val))
        numeric_val = max_val

    if numeric_val < 0:
        warnings.append("{} score {} is negative — set to 0".format(column_name, numeric_val))
        numeric_val = 0

    # Return as int using math.ceil
    import math
    return (int(math.ceil(numeric_val)), warnings)

def compute_derived_scores(student_scores, config=None):
    """Given a dict of {column: value}, compute Total CA, Grand Total, Grade, Remarks.
    Returns (derived_dict, warnings_list).
    derived_dict adds: Total CA, Grand Total, Grade, Remarks
    """
    if config is None:
        config = NIGERIAN_MARK_BOOK_CONFIG
    warnings = []
    result = dict(student_scores)  # Copy original

    # Collect CA values
    ca_sum = 0
    ca_count = 0
    has_any_ca = False
    for ca_name in config["ca_columns"]:
        val = student_scores.get(ca_name)
        if val is not None and str(val).strip() != '' and str(val).upper() != 'ABS':
            try:
                ca_sum += float(val)
                ca_count += 1
                has_any_ca = True
            except (ValueError, TypeError):
                pass

    # Compute Total CA = sum(CAs) / 2
    if has_any_ca:
        import math
        total_ca = int(math.ceil(ca_sum / 2.0))
        if total_ca > config["ca_total_max"]:
            warnings.append("Total CA {} exceeds max {} — capped".format(total_ca, config["ca_total_max"]))
            total_ca = config["ca_total_max"]
        result["Total CA"] = total_ca

        # Check if AI-provided Total CA differs
        ai_total_ca = student_scores.get("Total CA")
        if ai_total_ca is not None and str(ai_total_ca).strip() != '' and str(ai_total_ca).upper() != 'ABS':
            try:
                ai_val = float(ai_total_ca)
                if abs(ai_val - total_ca) > 0.5:
                    warnings.append("AI Total CA ({}) differs from computed ({})".format(ai_val, total_ca))
            except (ValueError, TypeError):
                pass
    else:
        total_ca = None

    # Get Exam score
    exam_val = student_scores.get("Exam")
    exam_numeric = None
    if exam_val is not None and str(exam_val).strip() != '' and str(exam_val).upper() != 'ABS':
        try:
            exam_numeric = float(exam_val)
        except (ValueError, TypeError):
            pass

    # Compute Grand Total = Total CA + Exam
    if total_ca is not None and exam_numeric is not None:
        import math
        grand_total = int(math.ceil(total_ca + exam_numeric))
        if grand_total > config["grand_total_max"]:
            warnings.append("Grand Total {} exceeds max {} — capped".format(grand_total, config["grand_total_max"]))
            grand_total = config["grand_total_max"]
        result["Grand Total"] = grand_total

        # Grade and Remarks
        for low, high, grade, remark in config["grade_map"]:
            if low <= grand_total <= high:
                result["Grade"] = grade
                result["Remarks"] = remark
                break
    elif total_ca is not None:
        result["Grand Total"] = ""
        result["Grade"] = ""
        result["Remarks"] = ""

    return (result, warnings)

def compute_term_averages(term_totals, current_term):
    """Compute cumulative term average.
    term_totals: {"1st Term": 72, "2nd Term": 68, "3rd Term": 75}
    current_term: "2nd Term" → (72 + 68) / 2 = 70
    current_term: "3rd Term" → (72 + 68 + 75) / 3 = 71.67
    Returns: (average, terms_included_count) or (None, 0) if no valid data
    """
    terms_order = ["1st Term", "2nd Term", "3rd Term"]
    if current_term not in terms_order:
        return (None, 0)
    end_idx = terms_order.index(current_term) + 1
    relevant_terms = terms_order[:end_idx]

    valid_totals = []
    for t in relevant_terms:
        val = term_totals.get(t)
        if val is not None and str(val).strip() != '' and str(val).upper() != 'ABS':
            try:
                valid_totals.append(float(val))
            except (ValueError, TypeError):
                pass

    if not valid_totals:
        return (None, 0)
    import math
    return (int(math.ceil(sum(valid_totals) / len(valid_totals))), len(valid_totals))

def get_grade_and_remark(score, config=None):
    """Get grade and remark for a numeric score."""
    if config is None:
        config = NIGERIAN_MARK_BOOK_CONFIG
    try:
        val = float(score)
    except (ValueError, TypeError):
        return ("", "")
    for low, high, grade, remark in config["grade_map"]:
        if low <= val <= high:
            return (grade, remark)
    return ("", "")

def format_position(rank):
    """Format a numeric rank into ordinal position string (1st, 2nd, 3rd, etc.)."""
    if pd.isna(rank): return ''
    r = int(rank)
    if 11 <= (r % 100) <= 13: return "{}th".format(r)
    suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(r % 10, 'th')
    return "{}{}".format(r, suffix)

# The prompt instructions for the AI model
SYSTEM_PROMPT = """
You are an expert OCR Assistant helping a Nigerian teacher grade test scripts.
I will provide you with images of handwritten test scripts. 
For EACH image, extract the following information:
1. Student Name (usually found at the top, e.g., "Name: ..."). Names may be long and written in ALL CAPS block letters (e.g., "ALARE OLUWAPELUMI OPEYEMI"). Combine first name, middle name, and surname into one field.
2. Student Class (usually found near the name, e.g., "Class: ..."). Classes might contain superscript letters (like SS1^Q or cursive); normalize this to standard text (e.g., "SS1Q").
3. Score (This is typically handwritten in **red ink**, often as a fraction like "8/10". Look carefully at the left margin — the score might be written VERY LARGE, spanning multiple lines, or circled. Always combine numerator and denominator if spread out.)
   - NIGERIAN MARK BOOK AWARENESS: Scores may be out of 10 (for CA tests) or 70 (for Exams). "8/10" means a CA score of 8. "52/70" means an Exam score of 52.
   - Extract ONLY the numerator (the student's actual mark).
   - If the score says "AB", "-", "/" or "NIL", this means ABSENT — return "ABS" as the score.
   - Handle fractional scores: 6½ = 6.5, 8½ = 8.5
   - If a number is crossed out and rewritten, use the CORRECTED (newer) value.
4. Confidence (How confident you are that the extracted data is correct: "High", "Medium", or "Low". Use "Low" when handwriting is very messy or fields are partially obscured.)

If an image is unreadable or you cannot find a specific field, return null for that field and set confidence to "Low".

Return exactly a JSON array of objects, one object for each image provided, in the exact same order as the images are provided.
Use the following JSON schema:
[
  {
    "name": "Extracted Name",
    "class": "Extracted Class",
    "score": "Extracted Score",
    "confidence": "High | Medium | Low"
  },
  ...
]

Return ONLY the raw JSON array. DO NOT wrap it in markdown block quotes like ```json ... ```.
"""


# Self-ping to keep Render free tier awake (pings every 14 min)
def _keep_alive():
    """Background thread that pings the app to prevent Render free-tier spin-down."""
    import urllib.request
    render_url = os.getenv("RENDER_EXTERNAL_URL")  # Auto-set by Render
    if not render_url:
        print("RENDER_EXTERNAL_URL not set — self-ping disabled (local dev).")
        return
    ping_url = render_url.rstrip("/") + "/health"
    print("Self-ping enabled: {} every 14 min".format(ping_url))
    while True:
        time.sleep(14 * 60)  # 14 minutes
        try:
            urllib.request.urlopen(ping_url, timeout=10)
            logger.debug("Self-ping OK")
        except Exception as e:
            logger.warning("Self-ping failed: {}".format(e))

_ping_thread = threading.Thread(target=_keep_alive, daemon=True)
_ping_thread.start()

# ═══════════════════════════════════════════════════════════════
#  GENERATED EXCEL FILE CLEANUP
#  Prevents disk exhaustion by removing stale generated spreadsheets.
# ═══════════════════════════════════════════════════════════════
EXCEL_MAX_AGE_SECONDS = 3600  # 1 hour
_PROTECTED_EXCEL_FILES = {
    "ActiveRoaster.xlsx",
    "Mathematics_1stTerm_SS1.xlsx",
    "Math_2ndTerm_SS1.xlsx",
    "Chemistry_SS1.xlsx",
    "test.xlsx",
}

def _cleanup_old_excel_files(max_age_seconds=None):
    """Remove generated .xlsx files older than max_age_seconds from the app root.
    Preserves protected files (ActiveRoaster, sample spreadsheets).
    Runs silently — never raises."""
    if max_age_seconds is None:
        max_age_seconds = EXCEL_MAX_AGE_SECONDS
    app_dir = os.path.dirname(os.path.abspath(__file__))
    now = time.time()
    removed = 0
    try:
        for f in glob.glob(os.path.join(app_dir, "*.xlsx")):
            basename = os.path.basename(f)
            if basename in _PROTECTED_EXCEL_FILES:
                continue
            try:
                age = now - os.path.getmtime(f)
                if age > max_age_seconds:
                    os.remove(f)
                    removed += 1
                    logger.info("Cleanup: removed stale file '{}' (age={:.0f}s)".format(basename, age))
            except OSError:
                pass  # File in use or already deleted
    except Exception as e:
        logger.warning("Cleanup sweep error: {}".format(e))
    return removed

@app.route('/health')
def health_check():
    """Health check endpoint for Render and monitoring."""
    try:
        # Quick DB check
        db.session.execute(text('SELECT 1'))
        return jsonify({"status": "healthy", "db": "ok", "ai_keys": len(API_KEYS)}), 200
    except Exception as e:
        logger.error("Health check failed: {}".format(e))
        return jsonify({"status": "unhealthy", "error": str(e)}), 503

@app.route('/')
def index():
    # Attempt to bust cache so user sees the new loading overlay
    response = make_response(render_template('index.html', v=int(time.time())))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.route('/api/recent-sessions', methods=['GET'])
def recent_sessions():
    """Returns classes with student counts and existing assessment types for the landing page."""
    try:
        classes = ClassModel.query.order_by(ClassModel.name).all()
        sessions = []
        seen_names = set()
        for c in classes:
            normalized = c.name.lower().replace(" ", "")
            if normalized in seen_names:
                continue
            seen_names.add(normalized)
            
            student_count = StudentModel.query.filter_by(class_id=c.id).count()
            if student_count == 0:
                continue
                
            # Find unique subjects and assessment types for this class
            students = StudentModel.query.filter_by(class_id=c.id).all()
            student_ids = [s.id for s in students]
            scores = ScoreModel.query.filter(ScoreModel.student_id.in_(student_ids)).all()
            
            subjects = {}
            for score in scores:
                subj = score.subject_name or 'Uncategorized'
                if subj not in subjects:
                    subjects[subj] = set()
                subjects[subj].add(score.assessment_type)
            
            if subjects:
                for subj, assessments in subjects.items():
                    sessions.append({
                        "class_id": c.id,
                        "class_name": c.name,
                        "subject": subj,
                        "student_count": student_count,
                        "existing_assessments": sorted(list(assessments)),
                        "suggested_next": suggest_next_assessment(sorted(list(assessments)))
                    })
            else:
                sessions.append({
                    "class_id": c.id,
                    "class_name": c.name,
                    "subject": None,
                    "student_count": student_count,
                    "existing_assessments": [],
                    "suggested_next": "1st CA"
                })
        
        return jsonify(sessions), 200
    except Exception as e:
        print("Error fetching recent sessions: {}".format(e))
        return jsonify([]), 200

def suggest_next_assessment(existing):
    """Suggest the next logical assessment type based on what already exists.
    Uses the full Nigerian mark book assessment order.
    """
    order = NIGERIAN_MARK_BOOK_CONFIG["assessment_order"]
    # Skip computed columns (Total CA, Grand Total) — those are auto-calculated
    scannable = [a for a in order if a not in ["Total CA", "Grand Total"]]
    for a in scannable:
        if a not in existing:
            return a
    return "Complete ✓"

@app.route('/api/classes', methods=['GET', 'POST'])
def handle_classes():
    if request.method == 'GET':
        classes = ClassModel.query.order_by(ClassModel.name).all()
        # seen = set() # Original line
        unique_classes = []
        seen_names = set() # Added as per instruction
        for c in classes:
            normalized_name = c.name.lower().replace(" ", "") # Modified as per instruction
            if normalized_name not in seen_names: # Modified as per instruction
                seen_names.add(normalized_name) # Modified as per instruction
                student_count = StudentModel.query.filter_by(class_id=c.id).count()
                unique_classes.append({"id": c.id, "name": c.name, "student_count": student_count})
        return jsonify(unique_classes), 200
        
    if request.method == 'POST':
        # Support both JSON (fallback) and Form Data (new UI)
        if request.is_json:
            data = request.get_json()
            raw_name = data.get('name', '').strip()
            names_text = data.get('names_text', '')
            subject_name = data.get('subject', 'Uncategorized')
            if not subject_name.strip(): subject_name = 'Uncategorized'
            file = None
        else:
            raw_name = request.form.get('name', '').strip()
            names_text = request.form.get('names_text', '')
            subject_name = request.form.get('subject', 'Uncategorized')
            if not subject_name.strip(): subject_name = 'Uncategorized'
            file = request.files.get('file')

        if not raw_name:
            return jsonify({"error": "Class name is required"}), 400

        # Normalize class name: uppercase, split letters from digits, add space (e.g. 'ss1q' -> 'SS 1Q')
        c_cleaned = re_mod.sub(r'[^A-Z0-9]', '', raw_name.upper())
        match = re_mod.match(r'([A-Z]+)(\d+.*)', c_cleaned)
        normalized_name = "{} {}".format(match.group(1), match.group(2)) if match else raw_name.strip().upper()
            
        print("Upserting class: {} (normalized: {})".format(raw_name, normalized_name))
            
        # Get or Create class (Upsert mechanism)
        c = ClassModel.query.filter(func.lower(ClassModel.name) == normalized_name.lower()).first()
        if not c:
            c = ClassModel(name=normalized_name)
            db.session.add(c)
            db.session.commit()
        
        names_added = 0
        scores_imported = 0
        
        # Process Pasted Text
        if names_text:
            lines = [line.strip().title() for line in names_text.split('\n') if line.strip()]
            for name in lines:
                if not StudentModel.query.filter_by(class_id=c.id, name=name).first():
                    db.session.add(StudentModel(class_id=c.id, name=name))
                    names_added += 1
            db.session.commit()
            
        # Process Uploaded File
        if file and file.filename:
            if file.filename.lower().endswith(('.txt', '.md')):
                content = file.read().decode('utf-8')
                lines = [line.strip().title() for line in content.split('\n') if line.strip()]
                for name in lines:
                    if not StudentModel.query.filter_by(class_id=c.id, name=name).first():
                        db.session.add(StudentModel(class_id=c.id, name=name))
                        names_added += 1
                db.session.commit()
            else:
                try:
                    if file.filename.lower().endswith('.csv'):
                        df = pd.read_csv(file)
                    else:
                        df = pd.read_excel(file)
                    
                    if 'Name' in df.columns:
                        for index, row in df.iterrows():
                            # Clean Name
                            name = str(row['Name']).strip().title()
                            if not name or name.lower() == 'nan' or name.lower() == 'none':
                                continue
                                
                            student = StudentModel.query.filter_by(class_id=c.id, name=name).first()
                            if not student:
                                student = StudentModel(class_id=c.id, name=name)
                                db.session.add(student)
                                db.session.commit() # Need ID for score associations
                                names_added += 1
                                
                            # Import Excel scores
                            for col in df.columns:
                                if col.strip() in ['Name', 'Total Score', 'Position', 'Rank'] or col.startswith('Unnamed'):
                                    continue
                                    
                                val = str(row[col]).strip()
                                if val and val.lower() != 'nan' and val.lower() != 'none':
                                    s_rec = ScoreModel.query.filter_by(
                                        student_id=student.id, 
                                        assessment_type=col.strip(), 
                                        subject_name=subject_name
                                    ).first()
                                    if s_rec:
                                        s_rec.score_value = val
                                    else:
                                        s_rec = ScoreModel(
                                            student_id=student.id, 
                                            score_value=val, 
                                            assessment_type=col.strip(), 
                                            subject_name=subject_name
                                        )
                                        db.session.add(s_rec)
                                        scores_imported += 1
                                        
                        db.session.commit()
                    else:
                        return jsonify({"error": "Excel/CSV file MUST contain a 'Name' column header."}), 400
                except Exception as e:
                    return jsonify({"error": "Error parsing file: {}".format(str(e))}), 500
                    
        msg = "Class '{}' ready. Added {} new students.".format(raw_name, names_added)
        if scores_imported > 0:
            msg += " Imported {} absolute scores.".format(scores_imported)
        elif names_text and names_added == 0:
            msg = "Class '{}' recognized. All provided students were already enrolled (0 duplicates injected). You are good to go!".format(raw_name)
            
        return jsonify({
            "success": True, 
            "message": msg
        }), 201

@app.route('/api/students', methods=['GET', 'POST'])
def handle_students():
    if request.method == 'GET':
        class_id = request.args.get('class_id')
        class_name = request.args.get('class_name')
        
        if class_id:
            students = StudentModel.query.filter_by(class_id=class_id).order_by(StudentModel.name).all()
        elif class_name:
            c = ClassModel.query.filter(func.lower(ClassModel.name) == str(class_name).strip().lower()).first()
            if not c:
                return jsonify([]), 200
            students = StudentModel.query.filter_by(class_id=c.id).order_by(StudentModel.name).all()
        else:
            return jsonify({"error": "class_id or class_name is required"}), 400
            
        return jsonify([{"id": s.id, "name": s.name} for s in students]), 200
        
    if request.method == 'POST':
        data = request.json
        name = data.get('name', '').strip().title()
        class_id = data.get('class_id')
        
        if not name or not class_id:
            return jsonify({"error": "name and class_id are required"}), 400
            
        c = ClassModel.query.get(class_id)
        if not c:
            return jsonify({"error": "Class not found"}), 404
            
        # Check if student exists
        existing = StudentModel.query.filter_by(class_id=class_id, name=name).first()
        if existing:
            return jsonify({"error": "Student '{}' already exists in this class.".format(name), "student": {"id": existing.id, "name": existing.name}}), 409
            
        s = StudentModel(class_id=class_id, name=name)
        db.session.add(s)
        db.session.commit()
        
        return jsonify({"message": "Student added successfully.", "student": {"id": s.id, "name": s.name}}), 201

@app.route('/api/students/<int:student_id>', methods=['DELETE', 'PUT'])
def manage_student(student_id):
    student = StudentModel.query.get(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404
    
    if request.method == 'DELETE':
        name = student.name
        db.session.delete(student)
        db.session.commit()
        return jsonify({"message": "Removed '{}'.".format(name)}), 200
    
    if request.method == 'PUT':
        data = request.json
        new_name = data.get('name', '').strip().title()
        if not new_name:
            return jsonify({"error": "Name is required"}), 400
        old_name = student.name
        student.name = new_name
        db.session.commit()
        return jsonify({"message": "Renamed '{}' to '{}'.".format(old_name, new_name), "student": {"id": student.id, "name": student.name}}), 200

@app.route('/api/enrollments', methods=['GET', 'POST'])
def handle_enrollments():
    if request.method == 'GET':
        class_id = request.args.get('class_id')
        subject_name = request.args.get('subject_name')
        if not class_id or not subject_name:
            return jsonify({"error": "class_id and subject_name are required"}), 400
            
        students = StudentModel.query.filter_by(class_id=class_id).all()
        enrolled_ids = []
        for s in students:
            enr = EnrollmentModel.query.filter_by(student_id=s.id, subject_name=subject_name).first()
            if enr:
                enrolled_ids.append(s.id)
                
        return jsonify({"enrolled": enrolled_ids}), 200
        
    if request.method == 'POST':
        data = request.json
        class_id = data.get('class_id')
        subject_name = data.get('subject_name')
        student_ids = data.get('student_ids', [])
        
        if not class_id or not subject_name:
            return jsonify({"error": "class_id and subject_name are required"}), 400
            
        # Clear existing enrollments for this class+subject
        students = StudentModel.query.filter_by(class_id=class_id).all()
        for s in students:
            EnrollmentModel.query.filter_by(student_id=s.id, subject_name=subject_name).delete()
            
        # Add new enrollments
        for sid in student_ids:
            enr = EnrollmentModel(student_id=sid, subject_name=subject_name)
            db.session.add(enr)
            
        db.session.commit()
        return jsonify({"message": "Enrollments updated successfully"}), 200

@app.route('/upload-scoresheet', methods=['POST'])
def upload_scoresheet():
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({"error": "No image provided"}), 400
            
        img_b64 = data['image']
        target_class = data.get('targetClass', '').strip()
        
        # Pull known names for this class
        known_names = []
        if target_class:
            try:
                c = ClassModel.query.filter(func.lower(ClassModel.name) == target_class.lower()).first()
                if c:
                    students = StudentModel.query.filter_by(class_id=c.id).all()
                    known_names = [s.name for s in students]
            except Exception as e:
                print("Error fetching known names: {}".format(e))
                
        # Engineer the Prompt per User Request
        roster_context = ""
        if known_names:
            roster_context = "\nCRITICAL ROSTER MAPPING: The students in this class are exactly: {}. You MUST map the extracted handwritten/typed names to exactly match a name from this list whenever visually possible. Do not invent alternate spellings if a direct visual resemblance exists in this roster.".format(known_names)
            
        system_prompt = """
You are an expert OCR Assistant helping a teacher digitize an entire grading score sheet.
I will provide you with an image of a score sheet (which may be handwritten or typed).
CRITICAL CONTEXT: The teacher has noted that some score sheets contain MULTIPLE columns of distinct grades for each student (e.g., a "1st CA" column, a "2nd CA" column, and an "Exam" column all on the same paper).

Your task is to:
1. Identify the table of students and extract ALL distinct score columns present.
2. Analyze the headers, title, or surrounding context to determine the precise names for these 'Assessment Types' (e.g., 1st CA, 2nd CA, Final Exam).
3. If only one column exists, infer its assessment type from context or label it "Score".

{}

Return EXACTLY a JSON object with two keys:
1. "assessment_types_found": A JSON array of strings listing the distinct assessment columns detected.
2. "records": A JSON array of objects. Each object must have a "name" string, and a "scores" dictionary mapping each detected assessment type string to its corresponding score value.

Use the following JSON schema:
{{
  "assessment_types_found": ["1st CA", "2nd CA", "Exam"],
  "records": [
    {{ 
       "name": "...", 
       "scores": {{ "1st CA": "...", "2nd CA": "...", "Exam": "..." }}
    }},
    {{ 
       "name": "...", 
       "scores": {{ "1st CA": "...", "2nd CA": "...", "Exam": "..." }}
    }}
  ]
}}

Return ONLY the raw JSON object. DO NOT wrap it in markdown block quotes like ```json ... ```.
""".format(roster_context)

        # Process with AI model — with retry + key rotation for rate limits
        raw_text = None
        for attempt in range(len(API_KEYS) if API_KEYS else 3):
            try:
                model = genai.GenerativeModel('gemini-2.5-flash')
                contents = [system_prompt, {"mime_type": "image/jpeg", "data": img_b64}]
                response = model.generate_content(contents)
                raw_text = response.text.strip()
                break
            except Exception as model_err:
                err_str = str(model_err).lower()
                if 'quota' in err_str or 'rate' in err_str or '429' in err_str or 'resource' in err_str:
                    rotate_api_key()
                    time.sleep(min(3 * (attempt + 1), 10))
                else:
                    raise
        if not raw_text:
            raise Exception('All API keys exhausted')
        
        if raw_text.startswith("```json"):
            raw_text = raw_text[7:]
        if raw_text.endswith("```"):
            raw_text = raw_text[:-3]
            
        result_json = json.loads(raw_text.strip())
        return jsonify(result_json), 200
        
    except Exception as e:
        print("Error processing score sheet: {}".format(e))
        return jsonify({"error": str(e)}), 500

@app.route('/upload-batch', methods=['POST'])
def upload_batch():
    try:
        data = request.json
        if not data or 'images' not in data:
            return jsonify({"error": "No images provided"}), 400
        
        images_base64 = data['images']
        target_class = data.get('targetClass', '').strip()
        target_classes = data.get('targetClasses', [target_class] if target_class else [])
        smart_instruction = data.get('smartInstruction', '').strip()
        
        if not images_base64:
             return jsonify({"error": "Empty images list"}), 400

        print("Received a batch of {} images for processing...".format(len(images_base64)))
        if smart_instruction:
            print("Smart Instruction Applied: {}".format(smart_instruction))
        
        # Build roster pool from ALL selected classes (3-Layer Routing)
        known_names_text = ""
        known_names = []
        class_rosters = {}  # {class_name: [student_names]}
        
        for tc in (target_classes if target_classes else [target_class]):
            if not tc:
                continue
            try:
                c = ClassModel.query.filter(func.lower(ClassModel.name) == tc.lower()).first()
                if c:
                    students = StudentModel.query.filter_by(class_id=c.id).all()
                    names = [s.name for s in students]
                    class_rosters[tc] = names
                    known_names.extend(names)
            except Exception as e:
                print("Error loading roster for {}: {}".format(tc, e))
        
        if known_names:
            known_names = list(set(known_names))  # Deduplicate
            known_names_text = "\n\nCRITICAL INSTRUCTION: You are grading papers for class(es) '{}'. Here is the authoritative list of known student names across all classes: {}. If the handwritten name on the paper resembles any of these, you MUST output the exact spelling from this list. Do not invent new names.".format(
                ', '.join(target_classes if target_classes else [target_class]),
                known_names
            )
        
        # Prepare contents for AI model
        dynamic_prompt = SYSTEM_PROMPT + known_names_text

        if smart_instruction:
            dynamic_prompt += "\n\n--- TEACHER'S CUSTOM SMART INSTRUCTION ---\n{}\n--- END OF CUSTOM INSTRUCTION ---\nYou MUST strictly obey the above manual instruction given by the teacher when processing these images and finalizing the output JSON.".format(smart_instruction)
        
        # Concurrent processing: split images into chunks and process in parallel
        CHUNK_SIZE = 5  # Increased from 3 - Gemini handles 5 images well, fewer API calls
        
        # Attach global index to each image
        indexed_images = list(enumerate(images_base64))
        
        # Split images into chunks
        image_chunks = []
        for i in range(0, len(indexed_images), CHUNK_SIZE):
            image_chunks.append(indexed_images[i:i + CHUNK_SIZE])
        
        def process_chunk(chunk_indexed_images):
            """Process a chunk of images through the AI model."""
            contents = [dynamic_prompt]
            for idx, img_b64 in chunk_indexed_images:
                if 'base64,' in img_b64:
                    img_b64 = img_b64.split('base64,')[1]
                contents.append({
                    "mime_type": "image/jpeg",
                    "data": img_b64
                })
            # Retry with key rotation on rate limit
            # Scale retries to number of keys (at least 3, up to keys * 2)
            max_retries = max(3, len(API_KEYS) * 2) if API_KEYS else 3
            backoff_times = [3, 5, 8, 12, 15]
            response = None
            model = genai.GenerativeModel('gemini-2.5-flash')
            for attempt in range(max_retries):
                try:
                    response = model.generate_content(contents)
                    break
                except Exception as retry_err:
                    err_str = str(retry_err).lower()
                    if 'quota' in err_str or 'rate' in err_str or '429' in err_str or 'resource' in err_str:
                        print("Rate limit hit on attempt {} — rotating key...".format(attempt + 1))
                        rotate_api_key()
                        # Re-create model with the newly configured key
                        model = genai.GenerativeModel('gemini-2.5-flash')
                        if attempt < max_retries - 1:
                            wait = backoff_times[min(attempt, len(backoff_times) - 1)]
                            time.sleep(wait)
                    else:
                        raise retry_err
            if not response:
                raise Exception("All API keys exhausted for this chunk after {} retries".format(max_retries))
            response_text = response.text.strip()
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            try:
                results = json.loads(response_text.strip())
            except Exception as e:
                print("Error parsing JSON from model: {}".format(e))
                results = []
                
            # Map back the global index to the result
            paired_results = []
            for i, res in enumerate(results):
                if i < len(chunk_indexed_images):
                    global_idx = chunk_indexed_images[i][0]
                    
                    # Clean the score (e.g., "8/10" -> "8")
                    raw_score = str(res.get('score', '')).strip()
                    if raw_score and '/' in raw_score:
                        res['score'] = raw_score.split('/')[0].strip()
                        
                    name = str(res.get('name', '')).strip().title()
                    confidence = str(res.get('confidence', 'high')).lower()
                    
                    # === 3-LAYER CLASS ROUTING ===
                    # Layer 1: OCR - try to match AI-extracted class
                    extracted_class = str(res.get('class', '')).strip().upper()
                    matched_class = None
                    
                    if extracted_class:
                        ec_cleaned = re_mod.sub(r'[^A-Z0-9]', '', extracted_class)
                        for tc in target_classes:
                            tc_cleaned = re_mod.sub(r'[^A-Z0-9]', '', tc.upper())
                            if ec_cleaned == tc_cleaned:
                                matched_class = tc
                                break
                    
                    # Layer 2: Roster lookup - find which class this student is in
                    if not matched_class and name and class_rosters:
                        for class_name, roster in class_rosters.items():
                            if roster:
                                best = process.extractOne(name, roster, scorer=fuzz.token_set_ratio)
                                if best and best[1] >= 85:
                                    matched_class = class_name
                                    res['name'] = best[0]  # Also fix name spelling
                                    break
                    
                    # Layer 3: Fallback to primary target class
                    if matched_class:
                        res['class'] = matched_class
                    elif target_class:
                        res['class'] = target_class

                    if name and known_names:
                        if confidence in ['low', 'medium'] or name not in known_names:
                            best_matches = process.extract(name, known_names, scorer=fuzz.token_set_ratio, limit=3)
                            
                            # Smart auto-correction if the top match is very high confidence and distinct
                            if best_matches and best_matches[0][1] >= 85:
                                # Check for ambiguity string ties
                                if len(best_matches) == 1 or best_matches[0][1] > best_matches[1][1] + 5:
                                    res['name'] = best_matches[0][0]
                                    res['needs_resolution'] = False
                                    res['fuzzy_matches'] = []
                                else:
                                    res['needs_resolution'] = True
                                    res['fuzzy_matches'] = [(m[0], m[1]) for m in best_matches]
                            elif best_matches:
                                res['needs_resolution'] = True
                                res['fuzzy_matches'] = [(m[0], m[1]) for m in best_matches]
                        
                    paired_results.append({"index": global_idx, "result": res})
            return paired_results
        
        @stream_with_context
        def generate():
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {executor.submit(process_chunk, chunk): chunk for chunk in image_chunks}
                for future in as_completed(futures):
                    try:
                        chunk_res = future.result()
                        for r in chunk_res:
                            yield "data: {}\n\n".format(json.dumps(r))
                    except Exception as exc:
                        print('Chunk generated an exception: {}'.format(exc))
                        yield "data: {}\n\n".format(json.dumps({"error": str(exc)}))
            yield "data: [DONE]\n\n"
        
        return Response(generate(), mimetype='text/event-stream')

    except Exception as e:
        print("Error processing batch: {}".format(e))
        return jsonify({"error": str(e)}), 500


def parse_class_level(class_name):
    """Smart class parser: extracts level and arm from any Nigerian school format.
    Handles: SS 1Q, JSS2A, ss1q, S.S 1 Q, J.S.S. 1B, SSS 3 Gold, Primary 4A, etc.
    Returns: {"level": "SS1", "arm": "Q", "normalized": "SS 1Q"}
    """
    raw = str(class_name).strip()
    # Step 1: Remove dots and extra whitespace, uppercase
    cleaned = re_mod.sub(r'\.', '', raw).upper().strip()
    cleaned = re_mod.sub(r'\s+', ' ', cleaned)
    
    # Step 2: Match pattern — (letters)(optional space)(digits)(optional space)(arm)
    # Handles: JSS 2A, SS1Q, SSS 3 Gold, PRIMARY 4A
    match = re_mod.match(r'^([A-Z]+)\s*(\d+)\s*(.*)$', cleaned)
    if match:
        prefix = match.group(1)  # SS, JSS, SSS, PRIMARY, PRI
        number = match.group(2)  # 1, 2, 3
        arm = match.group(3).strip()  # Q, A, Gold, or empty
        
        level = "{}{}".format(prefix, number)  # SS1, JSS2, SSS3
        normalized = "{} {}{}".format(prefix, number, arm)  # SS 1Q, JSS 2A
        
        return {"level": level, "arm": arm or "_default", "normalized": normalized}
    
    # Step 3: Fallback — treat entire name as level with no arm
    return {"level": cleaned or "UNKNOWN", "arm": "_default", "normalized": raw}


@app.route('/export-excel', methods=['POST'])
def export_excel():
    """Generates Excel from scanned results. Handles multi-term merge and standard formatting."""
    # Housekeeping: remove stale generated files to prevent disk exhaustion
    _cleanup_old_excel_files()
    try:
        data = request.json
        if not data or 'results' not in data:
            return jsonify({"error": "No results provided for export"}), 400
            
        results = data['results']
        assessment_type = data.get('assessmentType', 'Score').strip()
        subject_name = data.get('subjectType', data.get('subjectName', '')).strip()
        term = data.get('term', '1st Term').strip()  # 1st Term, 2nd Term, 3rd Term
        existing_records = data.get('existingRecords', None)
        
        if not assessment_type:
            assessment_type = 'Score'
        if not subject_name or subject_name.lower() == 'uncategorized':
            subject_name = data.get('subjectName', data.get('subjectType', '')).strip()
        if not subject_name:
            subject_name = 'General'
            
        if not results and not existing_records:
             return jsonify({"error": "No data to export"}), 400

        subject_mode = data.get('subjectMode', 'general').strip().lower()
             
        # === BUILD & MERGE EXCEL DATA ===
        
        # 1. Standardize and process new scanned results by class
        new_scores_by_class = {}
        for r in results:
            name = str(r.get('name', '')).strip().title()
            if not name: continue
            
            c_raw = str(r.get('class', '')).strip().upper()
            c_cleaned = re_mod.sub(r'[^A-Z0-9]', '', c_raw)
            match = re_mod.match(r'([A-Z]+)(\d+.*)', c_cleaned)
            class_name = "{} {}".format(match.group(1), match.group(2)) if match else (c_raw or "Unknown Class")
            
            if class_name not in new_scores_by_class:
                new_scores_by_class[class_name] = {}
            new_scores_by_class[class_name][name] = str(r.get('score', '')).strip()
            
        # 2. Add existing records and merge with new
        merged_by_class = {} # {class_name: {name: {col: val}}}
        
        if existing_records and isinstance(existing_records, list):
            for r in existing_records:
                # Expecting records like: {'Name': 'John', 'Class': 'SS 1A', '1st CA': 8, ...}
                name = str(r.get('Name', '')).strip().title()
                clsz = str(r.get('Class', '')).strip()
                if not name or not clsz: continue
                
                if clsz not in merged_by_class:
                    merged_by_class[clsz] = {}
                
                # Copy existing data exactly as is (ignoring computed columns which we'll recalculate)
                row_data = {k: v for k, v in r.items() if k not in ["Position", "Rank"]}
                # Make sure name and class are uniform
                row_data["Name"] = name
                row_data["Class"] = clsz
                
                merged_by_class[clsz][name] = row_data

        # 3. Correct scanned names to roster names BEFORE merging
        #    This prevents duplicates from OCR spelling differences
        for class_name, students in new_scores_by_class.items():
            c = ClassModel.query.filter(func.lower(ClassModel.name) == class_name.lower()).first()
            if c:
                db_students = StudentModel.query.filter_by(class_id=c.id).all()
                roster_names = [s.name for s in db_students]
                if roster_names:
                    corrected = {}
                    claimed = set()
                    for name, score in students.items():
                        if name in roster_names:
                            corrected[name] = score
                            claimed.add(name)
                        else:
                            available = [n for n in roster_names if n not in claimed]
                            if not available:
                                available = roster_names
                            best = process.extractOne(name, available, scorer=fuzz.token_set_ratio)
                            if best and best[1] >= 75:
                                corrected[best[0]] = score
                                claimed.add(best[0])
                            else:
                                # Drop unmatched OCR names — only roster names belong in the Excel
                                print("[ROSTER] Dropped unrecognized name '{}' (no roster match)".format(name))
                    new_scores_by_class[class_name] = corrected

        # 4. Merge corrected scans into existing records
        for class_name, students in new_scores_by_class.items():
            if class_name not in merged_by_class:
                merged_by_class[class_name] = {}
                
            for name, score in students.items():
                if not score: continue
                
                # Fuzzy match name against existing merged records to prevent duplicates
                existing_names = list(merged_by_class[class_name].keys())
                target_name = name
                if existing_names:
                    best = process.extractOne(name, existing_names, scorer=fuzz.token_set_ratio)
                    if best and best[1] >= 85:
                        target_name = best[0]
                
                if target_name not in merged_by_class[class_name]:
                    merged_by_class[class_name][target_name] = {"Name": target_name, "Class": class_name}
                
                # Update the new assessment column (this will overwrite previous session's value IF they regrade the SAME assessment)
                merged_by_class[class_name][target_name][assessment_type] = score
        
        # === ROSTER PADDING (All Subjects) ===
        # Always ensure all students known to the database for this class are listed.
        for class_name in list(merged_by_class.keys()):
            c = ClassModel.query.filter(func.lower(ClassModel.name) == class_name.lower()).first()
            if c:
                db_students = StudentModel.query.filter_by(class_id=c.id).all()
                existing_names = list(merged_by_class[class_name].keys())
                
                for db_stu in db_students:
                    target_name = db_stu.name
                    found = False
                    if existing_names:
                        best = process.extractOne(target_name, existing_names, scorer=fuzz.token_set_ratio)
                        if best and best[1] >= 85:
                            found = True
                    if not found:
                        # Pad with missing student
                        merged_by_class[class_name][target_name] = {"Name": target_name, "Class": class_name}
        
        # === COMPUTE & GROUP BY CLASS LEVEL ===
        level_groups = {}  # {level: {arm: [results]}}
        config = NIGERIAN_MARK_BOOK_CONFIG
        
        for class_name, students in merged_by_class.items():
            parsed = parse_class_level(class_name)
            level = parsed["level"]
            arm = parsed["arm"]
            
            if level not in level_groups:
                level_groups[level] = {}
            if class_name not in level_groups[level]:
                level_groups[level][class_name] = []
                
            for name, row_data in students.items():
                # Extract and clean ONLY the standard configured columns for computation
                # Other columns (like 1st Term Total) just pass through
                ca_scores = {}
                for col in config["ca_columns"]:
                    val = row_data.get(col)
                    if val is not None and str(val).strip() != '' and str(val).upper() != 'ABS':
                        try:
                            # Handle potential fractions like "8/10"
                            c_val = str(val).split('/')[0] if '/' in str(val) else val
                            ca_scores[col] = float(c_val)
                        except: pass
                
                if 'Exam' in row_data:
                    val = row_data['Exam']
                    if val is not None and str(val).strip() != '' and str(val).upper() != 'ABS':
                        try:
                            c_val = str(val).split('/')[0] if '/' in str(val) else val
                            ca_scores['Exam'] = float(c_val)
                        except: pass
                
                # Recompute Total CA, Grand Total, Grade, Remarks
                if ca_scores:
                    derived, _ = compute_derived_scores(ca_scores)
                    for key in ['Total CA', 'Grand Total', 'Grade', 'Remarks']:
                        if key in derived:
                            row_data[key] = derived[key]
                else:
                    for key in ['Total CA', 'Grand Total', 'Grade', 'Remarks']:
                        if key in row_data:
                            del row_data[key] # clear out old computations if scores were removed
                            
                # Ensure it's in the group
                level_groups[level][class_name].append(row_data)

        # === GENERATE EXCEL FILES ===
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        all_sheets_summary = {}
        generated_files = {}  # {level: filepath}
        
        for level, classes_in_level in level_groups.items():
            safe_subject = re_mod.sub(r'[^A-Za-z0-9 ]', '', subject_name).strip()
            # File named with term
            safe_term = re_mod.sub(r'[^A-Za-z0-9 ]', '', term).strip()
            filename = "{}_{}_{}.xlsx".format(safe_subject or "Scores", safe_term.replace(' ', ''), level)
            filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
            generated_files[level] = {"path": filepath, "filename": filename}
            
            sheets_dict = {}
            
            for class_name, rows in classes_in_level.items():
                # Exactly 3 terms as per physical mark book — NO Annual tab
                all_terms = ["1st Term", "2nd Term", "3rd Term"]
                base_df = pd.DataFrame(rows)
                if not base_df.empty and 'Name' in base_df.columns:
                    base_df = base_df.sort_values(by='Name', key=lambda col: col.str.lower()).reset_index(drop=True)
                
                # Insert Serial Number column at the far left
                if not base_df.empty:
                    base_df.insert(0, 'S/N', range(1, len(base_df) + 1))
                
                standard_ca_cols = ['1st CA', '2nd CA', 'Open Day', 'Note Book', 'Assignment']
                
                # Build a lookup of previous term Grand Totals from existingRecords
                prev_term_totals = {}  # {student_name: {"1st Term": total, "2nd Term": total}}
                if existing_records and isinstance(existing_records, list):
                    for rec in existing_records:
                        rname = str(rec.get('Name', '')).strip().title()
                        if not rname:
                            continue
                        if rname not in prev_term_totals:
                            prev_term_totals[rname] = {}
                        # Check for previous term total columns
                        for prev_key in ['1st Term Total', '2nd Term Total', '3rd Term Total']:
                            val = rec.get(prev_key)
                            if val is not None and str(val).strip() != '':
                                try:
                                    prev_term_totals[rname][prev_key] = float(val)
                                except (ValueError, TypeError):
                                    pass
                        # Also check Grand Total as a fallback for the term
                        gt = rec.get('Grand Total')
                        if gt is not None and str(gt).strip() != '':
                            try:
                                prev_term_totals[rname]['_grand_total'] = float(gt)
                            except (ValueError, TypeError):
                                pass
                
                for t in all_terms:
                    sheet_name = "{} - {}".format(class_name[:20], t[:10])
                    
                    if t == term:
                        # Active term — use the populated data
                        df = base_df.copy()
                    else:
                        # Inactive term — blank template with student names
                        empty_rows = [{"Name": r.get('Name'), "Class": r.get('Class')} for r in rows]
                        df = pd.DataFrame(empty_rows)
                        # Add Serial Number column for inactive term sheets too
                        if not df.empty:
                            df.insert(0, 'S/N', range(1, len(df) + 1))
                    
                    # Build standard columns: CA1-5, Total CA, Exam, Grand Total, Grade, Remarks
                    for col in standard_ca_cols + ['Total CA', 'Exam', 'Grand Total', 'Grade', 'Remarks']:
                        if col not in df.columns:
                            df[col] = ''
                    
                    # === CUMULATIVE COLUMNS based on term ===
                    if t == "2nd Term":
                        # Add "1st Term Total" column pulled from previous data
                        first_totals = []
                        for _, row in df.iterrows():
                            sname = str(row.get('Name', '')).strip().title()
                            pt = prev_term_totals.get(sname, {})
                            val = pt.get('1st Term Total', pt.get('_grand_total', ''))
                            first_totals.append(val)
                        df['1st Term Total'] = first_totals
                        
                        # "1st & 2nd" = 1st Term Total + 2nd Term Grand Total
                        cumulative = []
                        for _, row in df.iterrows():
                            t1 = row.get('1st Term Total', '')
                            t2 = row.get('Grand Total', '')
                            try:
                                t1_val = float(t1) if t1 != '' else None
                                t2_val = float(t2) if t2 != '' else None
                                if t1_val is not None and t2_val is not None:
                                    cumulative.append(round(t1_val + t2_val, 1))
                                else:
                                    cumulative.append('')
                            except (ValueError, TypeError):
                                cumulative.append('')
                        df['1st & 2nd'] = cumulative
                        
                    elif t == "3rd Term":
                        # Add "1st Term Total" and "2nd Term Total" columns
                        first_totals = []
                        second_totals = []
                        for _, row in df.iterrows():
                            sname = str(row.get('Name', '')).strip().title()
                            pt = prev_term_totals.get(sname, {})
                            first_totals.append(pt.get('1st Term Total', ''))
                            second_totals.append(pt.get('2nd Term Total', ''))
                        df['1st Term Total'] = first_totals
                        df['2nd Term Total'] = second_totals
                        
                        # "1st 2nd & 3rd" = sum of all 3 Grand Totals
                        cumulative = []
                        averages = []
                        for _, row in df.iterrows():
                            t1 = row.get('1st Term Total', '')
                            t2 = row.get('2nd Term Total', '')
                            t3 = row.get('Grand Total', '')
                            vals = []
                            for v in [t1, t2, t3]:
                                try:
                                    if v != '' and v is not None:
                                        vals.append(float(v))
                                except (ValueError, TypeError):
                                    pass
                            if vals:
                                total = round(sum(vals), 1)
                                avg = round(total / 3.0, 1)
                                cumulative.append(total)
                                averages.append(avg)
                            else:
                                cumulative.append('')
                                averages.append('')
                        df['1st 2nd & 3rd'] = cumulative
                        df['Average'] = averages
                    
                    # === BUILD FINAL COLUMN ORDER ===
                    final_cols = ['S/N', 'Name']
                    for col in standard_ca_cols:
                        final_cols.append(col)
                    final_cols.extend(['Total CA', 'Exam', 'Grand Total'])
                    
                    # Add cumulative columns in the right position
                    if t == "2nd Term":
                        final_cols.extend(['1st Term Total', '1st & 2nd'])
                    elif t == "3rd Term":
                        final_cols.extend(['1st Term Total', '2nd Term Total', '1st 2nd & 3rd', 'Average'])
                    
                    final_cols.extend(['Grade', 'Remarks'])
                    
                    # === RANKING ===
                    if t == "3rd Term" and 'Average' in df.columns:
                        # 3rd term: rank by Average
                        numeric_avg = pd.to_numeric(df['Average'], errors='coerce')
                        if numeric_avg.dropna().shape[0] > 0:
                            df['Rank'] = numeric_avg.rank(method='min', ascending=False)
                            df['Position'] = df['Rank'].apply(format_position)
                            df = df.drop(columns=['Rank'])
                        else:
                            df['Position'] = ''
                    elif t == term:
                        # Active term: rank by Grand Total or Total CA
                        rank_col = 'Grand Total' if 'Grand Total' in df.columns and df['Grand Total'].replace('', None).dropna().shape[0] > 0 else (
                                   'Total CA' if 'Total CA' in df.columns and df['Total CA'].replace('', None).dropna().shape[0] > 0 else None)
                        if rank_col:
                            numeric_scores = pd.to_numeric(df[rank_col], errors='coerce')
                            df['Rank'] = numeric_scores.rank(method='min', ascending=False)
                            df['Position'] = df['Rank'].apply(format_position)
                            df = df.drop(columns=['Rank'])
                        else:
                            df['Position'] = ''
                    else:
                        df['Position'] = ''
                    
                    final_cols.append('Position')
                    
                    # Ensure all columns exist
                    for col in final_cols:
                        if col not in df.columns:
                            df[col] = ''
                    
                    df = df[final_cols]
                    
                    import re
                    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

                    # Force nullable integer type to lose the .0 trailing decimals in output
                    for col in df.columns:
                        if col in ['Name', 'Class', 'Grade', 'Remarks', 'Position']:
                            # Remove illegal characters from strings to prevent openpyxl XML errors
                            df[col] = df[col].apply(lambda x: ILLEGAL_CHARACTERS_RE.sub('', str(x)) if pd.notna(x) else x).astype(str)
                        elif col not in ['S/N']:
                            # Round floats to nearest whole number so it can safely cast to Int64
                            df[col] = pd.to_numeric(df[col], errors='coerce').round().astype('Int64')
                    
                    sheets_dict[sheet_name] = df
            if not sheets_dict:
                continue
            
            # Write out
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for s_name, df_sheet in sheets_dict.items():
                    df_sheet.to_excel(writer, sheet_name=s_name, index=False, startrow=4)
                    worksheet = writer.sheets[s_name]
                    
                    worksheet.merge_cells('A1:K1')
                    title_cell = worksheet['A1']
                    # Parse the term from the sheet name (e.g. "SS 1Q - 1st Term" -> "1st Term")
                    sheet_term = s_name.split(' - ')[-1] if ' - ' in s_name else term
                    title_cell.value = "QSI SMART GRADER SCORESHEET - {}".format(sheet_term.upper())
                    title_cell.font = Font(bold=True, size=16)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    target_class = s_name.split(' - ')[0]
                    worksheet.merge_cells('A2:E2')
                    class_cell = worksheet['A2']
                    class_cell.value = "CLASS: {}".format(target_class)
                    class_cell.font = Font(bold=True, size=12)
                    
                    worksheet.merge_cells('A3:E3')
                    subj_cell = worksheet['A3']
                    subj_cell.value = "SUBJECT: {}".format(subject_name)
                    subj_cell.font = Font(bold=True, size=12)
                    
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
                    for col_idx in range(1, len(df_sheet.columns) + 1):
                        cell = worksheet.cell(row=5, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        
                        col_letter = get_column_letter(col_idx)
                        column_header = df_sheet.columns[col_idx - 1]
                        if column_header == 'Name':
                            worksheet.column_dimensions[col_letter].width = 30
                        elif column_header == 'Class':
                            worksheet.column_dimensions[col_letter].width = 12
                        elif column_header == 'Remarks':
                            worksheet.column_dimensions[col_letter].width = 15
                        else:
                            worksheet.column_dimensions[col_letter].width = 10
                            
                    for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row, min_col=3, max_col=worksheet.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
            
            for s_name, df_sheet in sheets_dict.items():
                raw_rows = df_sheet.to_dict(orient='records')
                # Clean NaNs and pd.NAs from rows so Flask jsonify doesn't output invalid JSON
                for r in raw_rows:
                    for k, v in r.items():
                        if pd.isna(v):
                            r[k] = ""
                            
                all_sheets_summary[s_name] = {
                    "columns": list(df_sheet.columns),
                    "rows": raw_rows,
                    "class": s_name.split(' - ')[0],
                    "subject": subject_name,
                    "level": level
                }
        
        if generated_files:
            import shutil
            first_level = list(generated_files.keys())[0]
            shutil.copy2(generated_files[first_level]["path"], WORKING_EXCEL_PATH)
        
        downloads = []
        for level, info in generated_files.items():
            safe_subject_url = re_mod.sub(r'[^A-Za-z0-9 ]', '', subject_name).strip() or 'Scores'
            safe_term_url = re_mod.sub(r'[^A-Za-z0-9 ]', '', term).strip()
            downloads.append({
                "level": level,
                "filename": info["filename"],
                "url": "/download-sheet?level={}&subject={}&term={}".format(level, safe_subject_url, safe_term_url)
            })

        return jsonify({
            "message": "Grades saved! {} file(s) ready for download.".format(len(downloads)),
            "sheets": all_sheets_summary,
            "downloads": downloads,
            "subject": subject_name
        }), 200

    except Exception as e:
        print("Excel export error: {}".format(e))
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/download-sheet', methods=['GET'])
def download_sheet():
    """Returns the compiled Excel file, optionally filtered by level."""
    level = request.args.get('level', '').strip()
    subject = request.args.get('subject', '').strip()
    term_param = request.args.get('term', '').strip()
    
    if level:
        # Look for level-specific file (must include term to match export_excel naming)
        safe_subject = re_mod.sub(r'[^A-Za-z0-9 ]', '', subject).strip() if subject else "Scores"
        safe_term = re_mod.sub(r'[^A-Za-z0-9 ]', '', term_param).strip().replace(' ', '') if term_param else ''
        if safe_term:
            filename = "{}_{}_{}.xlsx".format(safe_subject or "Scores", safe_term, level)
        else:
            filename = "{}_{}.xlsx".format(safe_subject or "Scores", level)
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        
        if os.path.exists(filepath):
            try:
                download_name = "{}_{}.xlsx".format(safe_subject or "Scores", level)
                response = make_response(send_file(
                    filepath, 
                    as_attachment=True, 
                    download_name=download_name,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ))
                response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
                response.headers["Pragma"] = "no-cache"
                response.headers["Expires"] = "0"
                return response
            except Exception as e:
                return jsonify({"error": str(e)}), 500
    
    # Fallback to working file
    if os.path.exists(WORKING_EXCEL_PATH):
        try:
            download_name = "{}_Scoresheet.xlsx".format(subject) if subject else "Scoresheet.xlsx"
            response = make_response(send_file(
                WORKING_EXCEL_PATH, 
                as_attachment=True, 
                download_name=download_name,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ))
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
            return response
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    else:
        return jsonify({"error": "No active sheet found. Start a batch first."}), 404

@app.route('/api/upload-excel-scorelist', methods=['POST'])
def upload_excel_scorelist():
    """Smart parser for Excel scorelists. Reads ALL term sheets for cumulative grading."""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({"error": "No file provided"}), 400
            
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({"error": "Only Excel or CSV files are supported"}), 400

        # Read Excel/CSV — smart header detection
        if file.filename.lower().endswith('.csv'):
            all_sheets = {"Sheet1": pd.read_csv(file)}
        else:
            # Read ALL sheets for multi-term support
            xf = pd.ExcelFile(file)
            all_sheets = {}
            for sn in xf.sheet_names:
                # Read without headers to dynamically find the row containing 'Name'
                df_raw = pd.read_excel(xf, sheet_name=sn, header=None)
                if df_raw.empty:
                    all_sheets[sn] = pd.DataFrame()
                    continue
                    
                header_row = 0
                for idx, row in df_raw.iterrows():
                    # Look for any cell containing 'name' (case-insensitive)
                    if any('name' in str(cell).lower().strip() for cell in row.values if pd.notna(cell)):
                        header_row = idx
                        break
                        
                # Re-read with correct header (use cached xf to avoid file pointer exhaustion)
                df_try = pd.read_excel(xf, sheet_name=sn, skiprows=header_row)
                all_sheets[sn] = df_try

        all_records = []
        term_data = {}  # {term: [{student records}]}
        detected_class = None
        detected_subject = None
        assessment_types = []

        for sheet_name, df in all_sheets.items():
            # Parse term from sheet name (e.g. "SS 1Q - 1st Term" -> "1st Term")
            sheet_term = None
            if ' - ' in sheet_name:
                parts = sheet_name.split(' - ')
                possible_term = parts[-1].strip()
                if 'term' in possible_term.lower():
                    sheet_term = possible_term

            # Detect Name Column
            name_col = None
            for col in df.columns:
                if 'name' in str(col).lower():
                    name_col = col
                    break
            if not name_col and not df.empty:
                name_col = df.columns[0]
            if not name_col:
                continue

            # Detect Class and Subject Columns
            class_col = None
            subj_col = None
            for col in df.columns:
                c_low = str(col).lower()
                if 'class' in c_low: class_col = col
                if 'subject' in c_low: subj_col = col

            # Detect Assessment types
            exclude_base = [name_col, class_col, subj_col, 'Total Score', 'Position', 'Rank', 'Total', 'S/N', 'S/N.', 'No', 'No.']
            exclude_lower = [str(x).lower().strip() for x in exclude_base if x]
            
            sheet_assessments = [
                str(col) for col in df.columns
                if str(col).lower().strip() not in exclude_lower
                and isinstance(col, str)           # skip integer/unnamed column indices
                and not str(col).startswith('Unnamed')
                and str(col).strip() != ''
            ]
            
            for a in sheet_assessments:
                if a not in assessment_types:
                    assessment_types.append(a)

            if not detected_class and class_col and not df.empty:
                detected_class = str(df[class_col].iloc[0])
            if not detected_subject and subj_col and not df.empty:
                detected_subject = str(df[subj_col].iloc[0])

            # Extract data from this sheet
            sheet_records = []
            for _, row in df.iterrows():
                name = str(row[name_col]).strip().title()
                if not name or name.lower() in ['nan', 'none']:
                    continue

                scores = {}
                for atype in sheet_assessments:
                    val = str(row[atype]).strip()
                    if val and val.lower() not in ['nan', 'none', '']:
                        scores[atype] = val

                # Skip rows with no scores at all (blank term sheets)
                if not scores:
                    continue

                # Guard against NaN class/subject values
                cls_val = str(row[class_col]).strip() if class_col else ''
                if cls_val.lower() in ['nan', 'none', '']:
                    cls_val = detected_class or ''
                subj_val = str(row[subj_col]).strip() if subj_col else ''
                if subj_val.lower() in ['nan', 'none', '']:
                    subj_val = detected_subject or ''
                r = {
                    "Name": name,
                    "Class": cls_val,
                    "Subject": subj_val,
                    "Term": sheet_term
                }
                # Flatten scores into the main record so /export-excel can read them naturally
                for k, v in scores.items():
                    r[k] = v
                
                sheet_records.append(r)
                all_records.append(r)

            if sheet_term and sheet_records:
                term_data[sheet_term] = sheet_records

        # DB Sync: Fuzzy-match uploaded names against roster to prevent phantom students.
        # Only genuinely new names (no close roster match) are added.
        if detected_class:
            c = ClassModel.query.filter(func.lower(ClassModel.name) == detected_class.lower()).first()
            if not c:
                c = ClassModel(name=detected_class.title())
                db.session.add(c)
                db.session.commit()

            roster_students = StudentModel.query.filter_by(class_id=c.id).all()
            roster_names = [s.name for s in roster_students]
            roster_names_lower = [n.lower() for n in roster_names]
            claimed = set()

            for r in all_records:
                s_name = r.get('Name', '').strip()
                if not s_name:
                    continue

                # Exact match — nothing to do
                if s_name.lower() in roster_names_lower:
                    continue

                # Fuzzy match against unclaimed roster names
                available = [n for n in roster_names if n not in claimed]
                if not available:
                    available = roster_names
                if available:
                    best = process.extractOne(s_name, available, scorer=fuzz.token_set_ratio)
                    if best and best[1] >= 75:
                        # Correct the record's name to the official roster version
                        logger.info("[UPLOAD ROSTER] Corrected '{}' -> '{}' (score={})".format(s_name, best[0], best[1]))
                        r['Name'] = best[0]
                        claimed.add(best[0])
                        continue

                # No match at all — genuinely new student, add to roster
                new_student = StudentModel(name=s_name.title(), class_id=c.id)
                db.session.add(new_student)
                roster_names.append(s_name.title())
                roster_names_lower.append(s_name.lower())
                logger.info("[UPLOAD ROSTER] Added new student '{}' (no roster match)".format(s_name))

            db.session.commit()

        return jsonify({
            "success": True,
            "total_students": len(all_records),
            "assessment_types_found": assessment_types,
            "records": all_records,
            "term_data": term_data,
            "detected_class": detected_class,
            "detected_subject": detected_subject,
            "filename": file.filename
        }), 200

    except Exception as e:
        err_detail = "{}: {}".format(type(e).__name__, str(e))
        logger.error("Excel scorelist upload error: {}".format(err_detail))
        return jsonify({"error": "Could not parse the Excel file. {}".format(err_detail)}), 500


@app.route('/api/extract-names', methods=['POST'])
def extract_names():
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({"error": "No image provided"}), 400
            
        img_b64 = data['image']
        target_class = data.get('targetClass', '').strip()
        
        known_names = []
        if target_class:
            c = ClassModel.query.filter(func.lower(ClassModel.name) == target_class.lower()).first()
            if c:
                students = StudentModel.query.filter_by(class_id=c.id).all()
                known_names = [s.name for s in students]
                
        system_prompt = """
You are an OCR assistant. I will provide an image of a handwritten or typed list of student names.
Extract the names you see.

Return EXACTLY a JSON array of strings, where each string is an extracted name.
For example: ["John Doe", "Jane Smith"]

Return ONLY the raw JSON array. DO NOT wrap it in markdown block quotes like `json ... `.
"""
        if known_names:
            system_prompt += "\nCRITICAL CONTEXT: Here is the authoritative list of known students in this class: {}. You MUST map the extracted handwritten names to exactly match a name from this list whenever visually possible.".format(known_names)
            
        # Process with AI model — with retry + key rotation for rate limits
        raw_text = None
        for attempt in range(len(API_KEYS) if 'API_KEYS' in dir() else 3):
            try:
                model = genai.GenerativeModel(
                    'gemini-2.5-flash',
                    generation_config=genai.GenerationConfig(
                        thinking_config=genai.types.ThinkingConfig(
                            thinking_budget=8192
                        )
                    )
                )
                contents = [system_prompt, {"mime_type": "image/jpeg", "data": img_b64}]
                response = model.generate_content(contents)
                # Extract text — handle thinking mode responses where thoughts may be separate parts
                raw_text = ''
                if hasattr(response, 'candidates') and response.candidates:
                    for part in response.candidates[0].content.parts:
                        if hasattr(part, 'thought') and part.thought:
                            continue  # Skip thought blocks, we only want the final JSON
                        if hasattr(part, 'text') and part.text:
                            raw_text += part.text
                if not raw_text:
                    raw_text = response.text.strip()
                raw_text = raw_text.strip()
                break
            except Exception as model_err:
                err_str = str(model_err).lower()
                if 'quota' in err_str or 'rate' in err_str or '429' in err_str or 'resource' in err_str:
                    rotate_api_key()
                    time.sleep(min(3 * (attempt + 1), 10))
                else:
                    raise
        if not raw_text:
            raise Exception('All API keys exhausted')
        
        if raw_text.startswith("```json"):
            raw_text = raw_text[7:]
        if raw_text.endswith("```"):
            raw_text = raw_text[:-3]
            
        result_json = json.loads(raw_text.strip())
        
        final_names = []
        for name in result_json:
            name_str = str(name).strip().title()
            if known_names:
                best_matches = process.extract(name_str, known_names, scorer=fuzz.token_set_ratio, limit=1)
                if best_matches and best_matches[0][1] >= 85:
                    final_names.append(best_matches[0][0])
                else:
                    final_names.append(name_str)
            else:
                final_names.append(name_str)
                
        return jsonify({"names": list(set(final_names))}), 200
        
    except Exception as e:
        print("Error extracting names: {}".format(e))
        return jsonify({"error": str(e)}), 500

@app.route('/api/ai-resolve', methods=['POST'])
def ai_resolve():
    """General-purpose AI resolver for sticky situations — called when the app hits ambiguity."""
    try:
        data = request.json
        situation = data.get('situation', '')
        context = data.get('context', {})
        
        resolver_prompts = {
            'unreadable_name': """A student's name could not be read from their test script.
Here is the context: {context}
The known student roster for this class is: {roster}
Based on any partial characters visible and the roster, suggest the most likely student name(s).
Return JSON: {{"suggestions": ["Name1", "Name2"], "confidence": "High|Medium|Low", "reasoning": "..."}}""",
            
            'weird_score': """A score was extracted but looks unusual: "{value}".
Common formats: "8", "8/10", "80%", "eight", "8 out of 10".
Normalize this to a simple integer score.
Return JSON: {{"normalized_score": 8, "original": "{value}", "reasoning": "..."}}""",
            
            'bulk_mismatch': """Multiple scanned names could not be matched to any student roster.
Unmatched names: {unmatched}
Available rosters: {rosters}
For each unmatched name, suggest the closest match from any roster, or mark as "new_student".
Return JSON: {{"matches": [{{"scanned": "...", "suggested": "...", "class": "...", "confidence": 0.85}}]}}""",
            
            'excel_format': """A teacher uploaded an Excel file but the format is unexpected.
Column headers found: {columns}
First 3 rows of data: {sample_rows}
Figure out which column is the student name, class, subject, and scores.
Return JSON: {{"name_col": "...", "class_col": "...", "subject_col": "...", "score_cols": ["..."], "header_row": 1, "reasoning": "..."}}""",
            
            'error_explain': """An error occurred in the grading app. Explain it in simple, friendly language for a teacher (not a developer).
Error: {error}
Context: What the teacher was doing: {action}
Return JSON: {{"friendly_message": "...", "suggestion": "...", "can_retry": true}}"""
        }
        
        template = resolver_prompts.get(situation, 
            'Analyze this situation and provide a helpful structured response: {context}')
        
        # Build the prompt with context
        prompt = template.format(**context) if context else template
        prompt += "\n\nReturn ONLY raw JSON. Do NOT wrap in markdown."
        
        # Process with AI — retry + rotate on rate limits
        raw_text = None
        for attempt in range(len(API_KEYS) if 'API_KEYS' in dir() else 3):
            try:
                model = genai.GenerativeModel('gemini-2.5-flash')
                response = model.generate_content(prompt)
                raw_text = response.text.strip()
                break
            except Exception as model_err:
                err_str = str(model_err).lower()
                if 'quota' in err_str or 'rate' in err_str or '429' in err_str or 'resource' in err_str:
                    rotate_api_key()
                    time.sleep(min(3 * (attempt + 1), 10))
                else:
                    raise
        if not raw_text:
            raise Exception('All API keys exhausted')
        
        if raw_text.startswith("```json"):
            raw_text = raw_text[7:]
        if raw_text.endswith("```"):
            raw_text = raw_text[:-3]
            
        result = json.loads(raw_text.strip())
        return jsonify({"success": True, "result": result}), 200
        
    except Exception as e:
        print("AI resolve error: {}".format(e))
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/move-student', methods=['POST'])
def move_student():
    """Move a student from one class to another."""
    try:
        data = request.json
        student_name = data.get('studentName', '').strip().title()
        from_class = data.get('fromClass', '').strip()
        to_class = data.get('toClass', '').strip()
        
        if not student_name or not from_class or not to_class:
            return jsonify({"error": "Missing student name, source class, or target class"}), 400
        
        # Find source class and student
        source = ClassModel.query.filter(func.lower(ClassModel.name) == from_class.lower()).first()
        if not source:
            return jsonify({"error": "Source class '{}' not found".format(from_class)}), 404
        
        student = StudentModel.query.filter_by(class_id=source.id, name=student_name).first()
        if not student:
            # Try fuzzy match
            existing = StudentModel.query.filter_by(class_id=source.id).all()
            names = [s.name for s in existing]
            if names:
                best = process.extractOne(student_name, names, scorer=fuzz.token_set_ratio)
                if best and best[1] >= 80:
                    student = StudentModel.query.filter_by(class_id=source.id, name=best[0]).first()
            if not student:
                return jsonify({"error": "Student '{}' not found in {}".format(student_name, from_class)}), 404
        
        # Ensure target class exists
        target = ClassModel.query.filter(func.lower(ClassModel.name) == to_class.lower()).first()
        if not target:
            target = ClassModel(name=to_class)
            db.session.add(target)
            db.session.commit()
        
        # Check for duplicate in target
        existing_in_target = StudentModel.query.filter_by(class_id=target.id, name=student.name).first()
        if existing_in_target:
            return jsonify({"error": "'{}' is already in {}".format(student.name, to_class)}), 400
        
        # Move the student
        student.class_id = target.id
        db.session.commit()
        
        return jsonify({
            "message": "Moved {} from {} to {} ✓".format(student.name, from_class, to_class),
            "success": True
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/api/assistant-scan-to-excel', methods=['POST'])
def assistant_scan_to_excel():
    """Receives an image + instruction, uses Vision AI to extract a table, returns an Excel file."""
    try:
        # Accept BOTH FormData (legacy) and JSON with base64 (iOS-safe)
        image_parts = []
        instruction = ''
        class_name = ''
        
        if request.is_json:
            # Safely parse JSON to prevent Flask default HTML error pages
            try:
                data = request.get_json(silent=True) or {}
            except Exception as json_err:
                print("JSON parsing error: {}".format(json_err))
                return jsonify({"error": "Failed to parse image data format."}), 400
            
            instruction = data.get('instruction', '').strip()
            class_name = data.get('class_name', '').strip()
            subject_name = data.get('subject_name', '').strip()
            assessment_type = data.get('assessment_type', '').strip()
            images_b64 = data.get('images_base64', [])
            
            if not images_b64:
                return jsonify({"error": "No images provided"}), 400
            
            for img in images_b64:
                try:
                    img_data = base64.b64decode(img.get('data', ''))
                    img_mime = img.get('mime_type', 'image/jpeg')
                    image_parts.append({"mime_type": img_mime, "data": img_data})
                except Exception as decode_err:
                    print("Image decode error: {}".format(decode_err))
        else:
            # FormData mode (legacy/desktop)
            if 'images' not in request.files:
                return jsonify({"error": "No images uploaded"}), 400
            
            instruction = request.form.get('instruction', '').strip()
            class_name = request.form.get('class_name', '').strip()
            subject_name = request.form.get('subject_name', '').strip()
            assessment_type = request.form.get('assessment_type', '').strip()
            files = request.files.getlist('images')
            
            for f in files:
                img_bytes = f.read()
                image_parts.append({"mime_type": f.content_type, "data": img_bytes})
        
        if not image_parts:
            return jsonify({"error": "No valid images received"}), 400
            
        # Hard validation: Do not allow scan without required context
        missing = []
        if not class_name: missing.append("class")
        if not subject_name: missing.append("subject")
        if not assessment_type: missing.append("term/assessment")
        
        if missing:
            return jsonify({"error": "Missing required information: {}. Please tell the assistant the missing info before scanning.".format(", ".join(missing))}), 400
            
        if not instruction:
            instruction = "extract all columns"
        
        # Build optional roster context for smarter OCR
        roster_context = ""
        roster_names = []
        matched_class = None
        if class_name:
            # Fuzzy class name lookup: "ss1s" should match "SS 1S"
            # First try exact (case-insensitive)
            c = ClassModel.query.filter(func.lower(ClassModel.name) == class_name.lower()).first()
            if not c:
                # Normalize: strip spaces/punctuation and compare
                normalized_input = re_mod.sub(r'[^a-zA-Z0-9]', '', class_name).lower()
                all_classes = ClassModel.query.all()
                for cls in all_classes:
                    normalized_db = re_mod.sub(r'[^a-zA-Z0-9]', '', cls.name).lower()
                    if normalized_db == normalized_input:
                        c = cls
                        break
                # If still not found, use thefuzz with a VERY strict threshold (classes differ by 1 letter often)
                if not c and all_classes:
                    class_names = [cls.name for cls in all_classes]
                    best = process.extractOne(class_name, class_names, scorer=fuzz.token_set_ratio)
                    if best and best[1] >= 95:
                        c = next((cls for cls in all_classes if cls.name == best[0]), None)
            
            if c:
                matched_class = c
                students = StudentModel.query.filter_by(class_id=c.id).all()
                if students:
                    roster_names = [s.name for s in students]
                    # Build a numbered roster for the AI to use as a lookup table
                    numbered_roster = '\n'.join(['  {}. {}'.format(i+1, name) for i, name in enumerate(roster_names)])
                    roster_context = """

===== MANDATORY NAME MATCHING =====
This image belongs to class '{class_name}'. There are {count} students enrolled in this class.
Here is the COMPLETE OFFICIAL ROSTER — every name the AI outputs MUST come from this list:

{numbered_roster}

RULES FOR NAME MATCHING:
- For EVERY row in the handwritten table, you MUST assign the "name" field to the CLOSEST matching name from the roster above.
- The handwriting may be messy, abbreviated, or misspelled. Use your best judgment to match each handwritten name to the correct roster entry.
- NEVER output a name that is NOT in the roster above. NEVER leave a name as empty string "".
- NEVER skip a student row. If you can see a row with scores, there MUST be a name for it.
- The number of rows you extract should approximately match the number of students in the roster ({count}).
- If you truly cannot determine which roster name a row belongs to, use your BEST GUESS from the roster — a wrong guess from the roster is better than an empty name or an invented name.
====================================""".format(class_name=c.name, count=len(roster_names), numbered_roster=numbered_roster)
        
        # No-roster fallback: class not in DB or no students yet
        if not roster_names and not roster_context:
            roster_context = """

===== NAME EXTRACTION (NO ROSTER AVAILABLE) =====
No class roster was provided for matching. You must read names directly from the handwriting.
RULES:
- Read EVERY name carefully, character by character. Nigerian names are often multi-part (e.g. "Abdulkareem Ihtimod Oyewumi").
- NEVER leave a name blank. Every row with scores MUST have a name.
- NEVER invent or guess names. Only write what you can actually SEE in the handwriting.
- If a name is partially readable, write the readable parts and use "?" for unclear characters.
- Pay attention to common Nigerian name patterns: Abdul-, Ade-, Ola-, Ayo-, Oba-, etc.
=================================================="""
        
        prompt = """
You are an expert OCR and data extraction AI specializing in Nigerian school record sheets. A teacher has uploaded image(s) of a handwritten document.

TEACHER'S INSTRUCTION: "{instruction}"{roster_context}

YOUR JOB: Read ALL the images and produce ONE combined JSON array of row objects representing the table data.

CRITICAL RULES:
1. **OUTPUT FORMAT**: Return ONLY a raw JSON array. Start with [ and end with ]. No markdown, no backticks, no explanations.
2. **COLUMN NAMING**: 
   - The student name column MUST always be keyed as "name" (lowercase).
   - If the teacher says "extract everything" or "all columns", auto-detect every column from the HEADER ROW and use readable names.
3. **ONLY EXTRACT WHAT IS PHYSICALLY WRITTEN**: 
   - Do NOT compute, calculate, or generate any values. Only extract what you can SEE written on the paper.
   - Do NOT add columns like "Grade", "Remarks", "Position", "Rank" etc. unless they are PHYSICALLY WRITTEN as a column on the sheet.
   - If you see columns like "Total CA", "Exam", "Grand Total" written on the sheet with handwritten values, extract them.
4. **COMMON COLUMN HEADERS** (detect even if handwritten messily):
   - "1st CA" / "1st Test" = First Continuous Assessment
   - "2nd CA" / "2nd Test" = Second Continuous Assessment  
   - "Open Day" / "Open" = Open Day score
   - "Note" / "NB" / "Note Book" = Notebook score
   - "Ass" / "Assig" / "Assignment" = Assignment score
   - "Attend" / "Attendance" = Attendance score
   - "Total CA" = Total Continuous Assessment
   - "Exam" = Examination score
   - "Grand Total" / "Total" = Final total score
5. **FRACTIONAL SCORES**: Convert fractions: 6½ → 6.5, 8½ → 8.5, 7½ → 7.5. If unsure, round to nearest 0.5.
6. **OVERWRITTEN VALUES**: If crossed out and rewritten, use the CORRECTED value.
7. **MISSING/UNREADABLE SCORES**: Mark completely unreadable or missing scores as exactly `null` (not "", not 0). This signals a gap for review.
8. **NAMES ARE MANDATORY**: Every single row MUST have a non-empty "name" field. NEVER output a row with an empty or missing name.
9. **SERIAL NUMBERS**: Do NOT include the S/N column unless specifically asked.
10. **MULTI-IMAGE**: If multiple images show pages of the SAME class, combine all rows into one array.
11. **NUMERIC VALUES**: All score values should be numbers (integers or decimals), NOT strings. Use 0 for zero, "" for missing.
12. **ROW-BY-ROW VERIFICATION**: For each row, verify: Is the name assigned? Are scores in correct columns? Do numbers make sense (CAs: 0-10, Exam: 0-70)?
""".format(instruction=instruction, roster_context=roster_context)

        # Call AI using centralized helper with retry + key rotation
        try:
            query_parts = [prompt] + image_parts
            raw_text = _call_gemini(AI_MODEL_PRIMARY, query_parts)
        except Exception as ai_err:
            err_str = str(ai_err).lower()
            logger.error("OCR scan AI failed after retries: {}".format(ai_err))
            if 'quota' in err_str or '429' in err_str or 'resource' in err_str:
                return jsonify({"error": "AI rate limit hit. Please wait about a minute and try again."}), 503
            return jsonify({"error": "AI Error: {}".format(str(ai_err)[:200])}), 503
            
        # Parse JSON
        raw_text = re_mod.sub(r'```json\n?', '', raw_text)
        raw_text = re_mod.sub(r'```\n?', '', raw_text)
        
        extracted_data = json.loads(raw_text)
        
        if not isinstance(extracted_data, list) or len(extracted_data) == 0:
            return jsonify({"error": "No valid data or table found in the image based on your instructions."}), 400
            
        # --- PASS 2: AI Smart Gap-Filling ---
        null_count = sum(1 for row in extracted_data for k, v in row.items() if v is None)
        if null_count > 0:
            logger.info("Found {} null cells. Triggering Pass 2 Smart Fill...".format(null_count))
            pass2_prompt = """
You are reviewing a partially extracted Nigerian school record sheet.
Here is what was extracted so far (null = unreadable):
{}

Using the ORIGINAL IMAGE, fill in ONLY the null cells.
Rules:
1. Use the image as the source of truth, not rigid math.
2. If you can SEE the value clearly in the handwriting, use it.
3. If you can logically INFER it from the row's other values and visible column totals, prefix it with ~ (e.g. "~8").
4. If you truly cannot determine it, leave it null.
5. Return ONLY the same JSON array structure with nulls filled. No explanation, no markdown.
""".format(json.dumps(extracted_data, indent=2))
            
            try:
                pass2_query = [pass2_prompt] + image_parts
                # Fallback model is fine for the simpler gap-filling task
                pass2_raw = _call_gemini(AI_MODEL_FALLBACK, pass2_query)
                pass2_raw = re_mod.sub(r'^```(?:json)?\s*', '', pass2_raw)
                pass2_raw = re_mod.sub(r'\s*```$', '', pass2_raw)
                pass2_data = json.loads(pass2_raw.strip())
                if isinstance(pass2_data, list) and len(pass2_data) == len(extracted_data):
                    extracted_data = pass2_data
                    logger.info("Pass 2 Smart Fill successful.")
            except Exception as p2_err:
                logger.warning("Pass 2 Smart Fill failed, falling back to Pass 1 data: {}".format(p2_err))
                
        # Clean up any remaining nulls back to "" for the frontend
        for row in extracted_data:
            for k, v in row.items():
                if v is None:
                    row[k] = ""
        
        # Post-OCR fuzzy name correction against the roster
        if roster_names:
            for row in extracted_data:
                ocr_name = str(row.get('name', '')).strip()
                if not ocr_name:
                    continue
                # Check if name already matches roster exactly
                if ocr_name in roster_names:
                    continue
                # Fuzzy match against roster
                best = process.extractOne(ocr_name, roster_names, scorer=fuzz.token_set_ratio)
                if best and best[1] >= 75:
                    row['name'] = best[0]  # Correct to official roster spelling
        
        # Get column names from the first row
        columns = list(extracted_data[0].keys()) if extracted_data else []
            
        matched_str = ""
        if 'roster_names' in locals() and roster_names and matched_class:
            matched_str = " (Matched roster: {})".format(matched_class.name)
        else:
            matched_str = " (No roster matched, used handwriting fallback)"

        # Return preview data instead of creating Excel immediately
        return jsonify({
            "success": True,
            "preview": True,
            "data": extracted_data,
            "columns": columns,
            "row_count": len(extracted_data),
            "message": "Found {} rows extracted{}. Review the data below — you can edit anything before building the Excel file.".format(len(extracted_data), matched_str)
        }), 200

    except json.JSONDecodeError as e:
        print("Assistant Image Scan JSON Error: {}".format(raw_text if raw_text else 'No response'))
        return jsonify({"error": "AI had trouble reading the image. Try a clearer photo or simpler instructions."}), 400
    except Exception as e:
        print("Assistant Image Scan Error: {}".format(e))
        return jsonify({"error": "Something went wrong: {}. Please try again.".format(str(e))}), 500


@app.route('/api/assistant-build-excel', methods=['POST'])
def assistant_build_excel():
    """Takes confirmed/edited preview data and builds the final Excel file."""
    # Housekeeping: remove stale generated files to prevent disk exhaustion
    _cleanup_old_excel_files()
    try:
        payload = request.json
        if not payload or 'data' not in payload:
            return jsonify({"error": "No data provided"}), 400
        
        extracted_data = payload['data']
        class_name = payload.get('class_name', '').strip()
        subject_name = payload.get('subject_name', '').strip()
        assessment_type = payload.get('assessment_type', '').strip()
        
        if not isinstance(extracted_data, list) or len(extracted_data) == 0:
            return jsonify({"error": "No data to build Excel from"}), 400
        
        df = pd.DataFrame(extracted_data)
        
        # Drop redundant 'Class' column if present
        df.drop(columns=['Class', 'class', 'CLASS'], errors='ignore', inplace=True)
        
        # --- PHASE 1: Correct OCR names to official roster names ---
        # --- PHASE 2: Pad with unscanned roster students ---
        # This ensures ALL names in the output come from the class roster.
        if class_name:
            c = ClassModel.query.filter(func.lower(ClassModel.name) == class_name.lower()).first()
            if c:
                db_students = StudentModel.query.filter_by(class_id=c.id).all()
                roster_names = [s.name for s in db_students]
                name_col = next((col for col in df.columns if str(col).lower() == 'name'), None)
                
                # PHASE 1: Correct every OCR name to the closest roster match
                # This prevents duplicates like "Abdulkareem I.O." vs "Abdulkareem Ihtimod Oyewumi"
                matched_roster_names = set()  # Track which roster names have been claimed
                if name_col and not df.empty and roster_names:
                    for idx in df.index:
                        ocr_name = str(df.at[idx, name_col]).strip()
                        if not ocr_name or ocr_name.lower() in ['nan', 'none']:
                            continue
                        # Check if this OCR name already exactly matches a roster name
                        if ocr_name in roster_names:
                            matched_roster_names.add(ocr_name)
                            continue
                        # Fuzzy match against UNCLAIMED roster names to avoid two OCR rows mapping to the same student
                        available_roster = [n for n in roster_names if n not in matched_roster_names]
                        if not available_roster:
                            available_roster = roster_names  # Fallback if all claimed
                        best = process.extractOne(ocr_name, available_roster, scorer=fuzz.token_set_ratio)
                        if best and best[1] >= 75:
                            df.at[idx, name_col] = best[0]  # Correct to official roster name
                            matched_roster_names.add(best[0])
                        # If no match at all (<75), drop the row — only roster names belong in the Excel
                        else:
                            print("[ROSTER] Dropped unrecognized name '{}' from preview (no roster match)".format(ocr_name))
                            df.drop(idx, inplace=True)
                
                # PHASE 2: Pad with roster students who had NO match in the scanned data
                if name_col and roster_names:
                    current_names = [str(n).strip() for n in df[name_col].tolist() if pd.notna(n) and str(n).strip()] if not df.empty else []
                    missing_students = []
                    for roster_name in roster_names:
                        # Check if this roster name is already in the DataFrame
                        found = False
                        if current_names:
                            best = process.extractOne(roster_name, current_names, scorer=fuzz.token_set_ratio)
                            if best and best[1] >= 85:
                                found = True
                        if not found:
                            row_dict = {name_col: roster_name}
                            for col in df.columns:
                                if col != name_col:
                                    row_dict[col] = ''
                            missing_students.append(row_dict)
                    
                    if missing_students:
                        df = pd.concat([df, pd.DataFrame(missing_students)], ignore_index=True)
        
        # Sort rows alphabetically by student name
        name_col = next((col for col in df.columns if str(col).lower() == 'name'), None)
        if name_col:
            df = df.sort_values(by=name_col, key=lambda col: col.str.lower()).reset_index(drop=True)
            
        # Add Serial Number (S/N) column at the far left
        df.insert(0, 'S/N', range(1, len(df) + 1))
        
        # Clean up any remaining '~' inferred markers if the teacher didn't edit them
        for col in df.columns:
            df[col] = df[col].apply(lambda x: str(x)[1:] if str(x).startswith('~') else x)
        
        # --- Normalize column names using grading engine ---
        rename_map = {}
        seen_normalized = set()
        for col in df.columns:
            normalized = normalize_column_name(col)
            base_normalized = normalized
            counter = 1
            while normalized in seen_normalized:
                normalized = "{} {}".format(base_normalized, counter)
                counter += 1
            seen_normalized.add(normalized)
            if normalized != str(col):
                rename_map[col] = normalized
        if rename_map:
            df = df.rename(columns=rename_map)
        
        # --- Validate and cap scores ---
        config = NIGERIAN_MARK_BOOK_CONFIG
        all_warnings = []
        score_cols = [c for c in df.columns if c in config["ca_columns"] or c in ["Exam", "Total CA", "Grand Total"]]
        for col in score_cols:
            for idx in df.index:
                val = df.at[idx, col]
                if val is not None and str(val).strip() != '':
                    cleaned, warns = validate_and_cap_score(col, val)
                    df.at[idx, col] = cleaned
                    all_warnings.extend(warns)
        
        # --- Auto-compute derived columns using grading engine ---
        ca_columns = [c for c in df.columns if c in config["ca_columns"]]
        has_enough_data = len(ca_columns) >= 2
        
        if has_enough_data:
            for idx in df.index:
                row_scores = {}
                for col in ca_columns:
                    val = df.at[idx, col]
                    if val is not None and str(val).strip() != '' and str(val).upper() != 'ABS':
                        try:
                            row_scores[col] = float(val)
                        except (ValueError, TypeError):
                            pass
                # Include Exam if present
                if 'Exam' in df.columns:
                    exam_val = df.at[idx, 'Exam']
                    if exam_val is not None and str(exam_val).strip() != '' and str(exam_val).upper() != 'ABS':
                        try:
                            row_scores['Exam'] = float(exam_val)
                        except (ValueError, TypeError):
                            pass
                
                if row_scores:
                    derived, warns = compute_derived_scores(row_scores)
                    all_warnings.extend(warns)
                    for key in ['Total CA', 'Grand Total', 'Grade', 'Remarks']:
                        if key in derived:
                            df.at[idx, key] = derived[key]
        
        # --- Add Position (ranking) ---
        if 'Grand Total' in df.columns:
            numeric_gt = pd.to_numeric(df['Grand Total'], errors='coerce')
            df['Position'] = numeric_gt.rank(method='min', ascending=False)
            df['Position'] = df['Position'].apply(format_position)
        
        # --- Reorder columns: S/N → name → CAs → Total CA → Exam → Grand Total → Grade → Remarks → Position ---
        desired_order = ['S/N', 'name'] + ca_columns
        for extra in ['Total CA', 'Exam', 'Grand Total', 'Grade', 'Remarks', 'Position']:
            if extra in df.columns:
                desired_order.append(extra)
        remaining = [c for c in df.columns if c not in desired_order]
        final_order = [c for c in desired_order if c in df.columns] + remaining
        df = df[final_order]
        
        # Build a descriptive filename
        name_parts = []
        if class_name:
            safe_class = re_mod.sub(r'[^A-Za-z0-9 ]', '', class_name).strip()
            name_parts.append(safe_class)
        if subject_name:
            safe_subject = re_mod.sub(r'[^A-Za-z0-9 ]', '', subject_name).strip()
            name_parts.append(safe_subject)
        if assessment_type:
            safe_assessment = re_mod.sub(r'[^A-Za-z0-9 ]', '', assessment_type).strip()
            name_parts.append(safe_assessment)
        
        if name_parts:
            output_filename = "{}_{}.xlsx".format('_'.join(name_parts), int(time.time()))
        else:
            output_filename = "Extracted_Data_{}.xlsx".format(int(time.time()))
        
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_filename)
        # --- Convert score columns to Nullable Integer 'Int64' to avoid trailing .0 ---
        import re
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        
        for col in df.columns:
            if col in ['S/N', 'name', 'Grade', 'Remarks', 'Position', 'Name', 'Class']:
                df[col] = df[col].apply(lambda x: ILLEGAL_CHARACTERS_RE.sub('', str(x)) if pd.notna(x) else x).astype(str)
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce').round().astype('Int64')
        
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            sheet_name_str = "Scores"
            if class_name and subject_name:
                sheet_name_str = "{} - {}".format(class_name[:15], subject_name[:15])
                
            df.to_excel(writer, sheet_name=sheet_name_str, index=False, startrow=4)
            worksheet = writer.sheets[sheet_name_str]
            
            # --- Emulate export_excel formatting ---
            worksheet.merge_cells('A1:K1')
            title_cell = worksheet['A1']
            title_cell.value = "QSI SMART GRADER SCORESHEET"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet.merge_cells('A2:E2')
            class_cell = worksheet['A2']
            class_cell.value = "Class: {}".format(class_name or "Unknown")
            class_cell.font = Font(bold=True)
            
            worksheet.merge_cells('F2:K2')
            term_cell = worksheet['F2']
            term_cell.value = "Term: {}".format("Active") # Assuming active term
            term_cell.font = Font(bold=True)
            term_cell.alignment = Alignment(horizontal='right')
            
            worksheet.merge_cells('A3:E3')
            subj_cell = worksheet['A3']
            subj_cell.value = "Subject: {}".format(subject_name or "Unknown")
            subj_cell.font = Font(bold=True)
            
            worksheet.merge_cells('F3:K3')
            assess_cell = worksheet['F3']
            assess_cell.value = "Assessment: {}".format(", ".join(assessment_type) if isinstance(assessment_type, list) else assessment_type)
            assess_cell.font = Font(bold=True)
            assess_cell.alignment = Alignment(horizontal='right')
            
            # Style headers
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            for cell in worksheet[5]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                
            # Auto-size columns
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                worksheet.column_dimensions[column].width = min(max_length + 2, 30)

        return jsonify({
            "success": True,
            "message": "Excel file ready! {} rows, {} columns.".format(len(df), len(df.columns)),
            "download_url": "/api/download-edited-excel?file={}".format(output_filename)
        }), 200
        
    except Exception as e:
        print("Build Excel Error: {}".format(e))
        return jsonify({"error": str(e)}), 500

@app.route('/api/assistant-edit-excel', methods=['POST'])
def assistant_edit_excel():
    """Receives an Excel file + natural language instruction, uses AI to apply edits, returns the modified file."""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        instruction = request.form.get('instruction', '').strip()
        if not instruction:
            return jsonify({"error": "No instruction provided"}), 400
        
        file = request.files['file']
        
        # Read the Excel into pandas
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # Build context for AI
        columns = list(df.columns)
        sample = df.head(5).to_string(index=False)
        row_count = len(df)
        
        prompt = """You are an Excel editing assistant. A teacher has uploaded a spreadsheet and wants you to edit it.

SPREADSHEET INFO:
- Columns: {columns}
- Total rows: {rows}
- First 5 rows:
{sample}

TEACHER'S INSTRUCTION: "{instruction}"

You must return a JSON object with an "edits" array. Each edit is one of:
1. {{"type": "update_cell", "row": 0, "column": "Score", "value": 85}} — update a specific cell (row is 0-indexed)
2. {{"type": "update_column", "column": "Score", "expression": "x + 5"}} — apply a math expression to an entire column (x = current value)
3. {{"type": "delete_rows", "condition_column": "Score", "condition": "< 40"}} — delete rows matching a condition
4. {{"type": "add_row", "data": {{"Name": "John", "Score": 80}}}} — add a new row
5. {{"type": "rename_column", "old_name": "Marks", "new_name": "Score"}} — rename a column
6. {{"type": "add_column", "column": "Assignment", "default_value": "0"}} - add a new column
7. {{"type": "confirm_column", "suspected_column": "Ass", "original_instruction": "{instruction}"}} — ASK FOR CONFIRMATION IF UNSURE.

RULES:
- If a teacher says "add 5 to [Column Name]", use "update_column" with expression "x + 5" for that specific column. 
- **CRITICAL SMARTNESS**: If that column does NOT exist, look at the Current Columns list. Is there a column with a very similar name or obvious abbreviation (e.g. they asked for "Assignment" but the column is "Ass" or "1st CA")? If YES, you MUST return ONLY ONE edit: "confirm_column" with your best guess. Do not proceed with adding columns or updating.
- If there is NO similar column, you MUST return TWO edits: first "add_column" to create it with default_value 0, then "update_column" to add 5 to it.
- x represents the current cell value. The math expression must use ONLY `x` (e.g. `x + 5`, `x * 10`).

Return ONLY raw JSON. Example:
{{"edits": [{{"type": "update_column", "column": "Assignment", "expression": "x + 5"}}], "summary": "Added 5 marks to Assignment column"}}
""".format(columns=columns, rows=row_count, sample=sample, instruction=instruction)
        
        # Call AI with retry
        raw_text = None
        for attempt in range(len(API_KEYS) if API_KEYS else 3):
            try:
                model = genai.GenerativeModel('gemini-2.5-flash')
                response = model.generate_content(prompt)
                raw_text = response.text.strip()
                break
            except Exception as model_err:
                err_str = str(model_err).lower()
                if 'quota' in err_str or 'rate' in err_str or '429' in err_str or 'resource' in err_str:
                    rotate_api_key()
                    time.sleep(min(3 * (attempt + 1), 10))
                else:
                    raise
        if not raw_text:
            raise Exception('All API keys exhausted')
        
        # Parse AI response
        if raw_text.startswith("```json"):
            raw_text = raw_text[7:]
        if raw_text.startswith("```"):
            raw_text = raw_text[3:]
        if raw_text.endswith("```"):
            raw_text = raw_text[:-3]
        
        ai_result = json.loads(raw_text.strip())
        edits = ai_result.get('edits', [])
        summary = ai_result.get('summary', 'Edits applied')
        
        # Apply edits to dataframe
        changes_made = 0
        needs_confirmation = None
        
        for edit in edits:
            edit_type = edit.get('type', '')
            
            if edit_type == 'confirm_column':
                needs_confirmation = {
                    "guess": edit.get('suspected_column'),
                    "original_instruction": edit.get('original_instruction')
                }
                break # Stop processing other edits if we need confirmation
                
            elif edit_type == 'update_cell':
                row = edit.get('row', 0)
                col = edit.get('column', '')
                if col in df.columns and 0 <= row < len(df):
                    df.at[row, col] = edit.get('value')
                    changes_made += 1
                    
            elif edit_type == 'add_column':
                col = edit.get('column', '')
                default_val = edit.get('default_value', '')
                if col and col not in df.columns:
                    # check if default_val is numeric string
                    if str(default_val).replace('.','',1).replace('-','',1).isdigit():
                        df[col] = float(default_val)
                    else:
                        df[col] = default_val
                    changes_made += 1
                    
            elif edit_type == 'update_column':
                col = edit.get('column', '')
                expr = edit.get('expression', '')
                if col in df.columns and expr:
                    try:
                        # Safe math evaluation — only allow basic arithmetic with x
                        def _safe_eval_expr(x_val, expression):
                            """Evaluate simple arithmetic expressions with x as the variable."""
                            import re as _re
                            # Only allow: digits, x, +, -, *, /, ., (, ), spaces
                            if not _re.match(r'^[\dx\.\+\-\*/\(\)\s]+$', expression):
                                raise ValueError("Unsafe expression: {}".format(expression))
                            # Replace 'x' with the actual value
                            safe_expr = expression.replace('x', str(float(x_val)))
                            return eval(safe_expr, {"__builtins__": {}}, {})
                        
                        df[col] = df[col].apply(lambda x: _safe_eval_expr(float(x) if str(x).replace('.','',1).replace('-','',1).isdigit() else 0, expr) if pd.notna(x) else x)
                        changes_made += len(df)
                    except Exception as eval_err:
                        print("Expression eval error: {}".format(eval_err))
                        
            elif edit_type == 'delete_rows':
                col = edit.get('condition_column', '')
                cond = edit.get('condition', '')
                if col in df.columns and cond:
                    try:
                        import re as _re
                        # Only allow: comparison operators, digits, x, basic math, spaces
                        if not _re.match(r'^[\dx\.\+\-\*/\(\)\s<>=!]+$', 'x ' + cond):
                            raise ValueError("Unsafe condition: {}".format(cond))
                        before_count = len(df)
                        def _safe_eval_cond(x_val, condition):
                            safe_expr = "x {}".format(condition).replace('x', str(float(x_val)))
                            return eval(safe_expr, {"__builtins__": {}}, {})
                        mask = df[col].apply(lambda x: _safe_eval_cond(float(x) if str(x).replace('.','',1).replace('-','',1).isdigit() else 0, cond) if pd.notna(x) else False)
                        df = df[~mask]
                        changes_made += before_count - len(df)
                    except Exception as eval_err:
                        print("Delete condition error: {}".format(eval_err))
                        
            elif edit_type == 'add_row':
                row_data = edit.get('data', {})
                if row_data:
                    df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)
                    changes_made += 1
                    
            elif edit_type == 'rename_column':
                old = edit.get('old_name', '')
                new = edit.get('new_name', '')
                if old in df.columns:
                    df = df.rename(columns={old: new})
                    changes_made += 1
        
        # Save modified file
        output_filename = "edited_{}".format(file.filename if file.filename else "output.xlsx")
        if not output_filename.endswith(('.xlsx', '.xls', '.csv')):
            output_filename += '.xlsx'
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_filename)
        
        if output_filename.endswith('.csv'):
            df.to_csv(output_path, index=False)
        else:
            df.to_excel(output_path, index=False, engine='openpyxl')
        
        
        if needs_confirmation:
            return jsonify({
                "success": True,
                "needs_confirmation": True,
                "guess": needs_confirmation["guess"],
                "original_instruction": needs_confirmation["original_instruction"],
                "message": "Did you mean the '{}' column?".format(needs_confirmation["guess"])
            }), 200
            
        if changes_made > 0:
            # Assuming WORKING_EXCEL_PATH is defined elsewhere, or we should use output_path
            # For now, using output_path as a placeholder if WORKING_EXCEL_PATH is not defined
            # If WORKING_EXCEL_PATH is meant to be a global variable, it should be defined.
            # For this edit, I'll assume it's a typo and should be output_path, or it's defined globally.
            # Given the context, it's likely a persistent working file.
            # If not defined, this line will cause an error. I'll keep it as is, assuming it's defined.
            # df.to_excel(WORKING_EXCEL_PATH, index=False) # Commented out as WORKING_EXCEL_PATH is not in provided context
            return jsonify({
                "success": True,
                "message": summary,
                "download_url": "/api/download-edited-excel?file={}".format(output_filename) # Changed to use output_filename
            }), 200
        else:
            return jsonify({
                "success": False,
                "error": "I understood the instruction, but it didn't result in any actual changes to the data."
            }), 200
            
    except Exception as e:
        print("Assistant edit Excel error: {}".format(e))
        return jsonify({"error": str(e)}), 500

@app.route('/api/download-edited-excel')
def download_edited_excel():
    """Download an edited Excel file, then schedule it for cleanup."""
    filename = request.args.get('file', '')
    if not filename or '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({"error": "Invalid filename"}), 400
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    if not os.path.exists(filepath):
        return jsonify({"error": "File not found"}), 404
    response = make_response(send_file(filepath, as_attachment=True, download_name=filename))
    # Clean up after serving — remove files older than 1 hour
    _cleanup_old_excel_files()
    return response


@app.route('/api/smart-assistant', methods=['POST'])
def smart_assistant():
    """Smart Assistant v3 — Full intelligence upgrade. Designed by XO.
    16 action types, proactive insights, anomaly detection, cross-class analysis.
    """
    try:
        data = request.json
        message = data.get('message', '')
        context = data.get('context', {})
        current_screen = data.get('currentScreen', 'landing')
        current_results = data.get('currentResults', [])  # Live scores from frontend
        session_info = data.get('sessionInfo', {})  # Classes graded, subject, etc.
        history = data.get('history', [])  # Conversation history
        images = data.get('images', [])  # Base64 images from frontend
        
        # Build SMART conversation context with action tracking
        conversation_context = ""
        last_proposed_action = None
        last_proposed_params = {}
        
        if history and len(history) > 1:
            conv_lines = []
            for h in history[:-1]:  # Exclude current message
                role = "Teacher" if h.get('role') == 'user' else "You (Assistant)"
                text = h.get('text', '')
                action = h.get('action', '')
                if text.startswith('[SYSTEM]'):
                    role = 'SYSTEM'
                    text = text[8:].strip()
                if action and action != 'none':
                    conv_lines.append("{}: {} [ACTION: {}]".format(role, text, action))
                    if role == "You (Assistant)":
                        last_proposed_action = action
                        last_proposed_params = h.get('params', {})
                else:
                    conv_lines.append("{}: {}".format(role, text))
            
            # Use sliding window for conversation history
            recent = conv_lines[-MAX_CONVERSATION_HISTORY:]
            conversation_context = "\nCONVERSATION HISTORY (maintain context!):\n{}\n".format("\n".join(recent))
            
            if last_proposed_action:
                conversation_context += "\nLAST PROPOSED ACTION: {} with params: {}\n".format(
                    last_proposed_action, json.dumps(last_proposed_params)
                )
                conversation_context += ">>> If teacher says 'yes'/'ok'/'do it'/'go ahead', RE-EXECUTE this action with SAME params! <<<\n"
        
        # Build RICH context from actual database — gives the AI real intelligence
        classes = ClassModel.query.order_by(ClassModel.name).all()
        class_data = {}
        grading_progress = []  # Proactive suggestions for the AI
        
        for c in classes:
            students = StudentModel.query.filter_by(class_id=c.id).order_by(StudentModel.name).all()
            scores = ScoreModel.query.filter(
                ScoreModel.student_id.in_([s.id for s in students])
            ).all() if students else []
            
            # Rich breakdown per class
            assessment_types = list(set(s.assessment_type for s in scores)) if scores else []
            subjects = list(set(s.subject_name for s in scores)) if scores else []
            terms = list(set(s.term for s in scores)) if scores else []
            
            # Per-subject grading completion check
            subject_progress = {}
            for subj in subjects:
                subj_scores = [s for s in scores if s.subject_name == subj]
                subj_assessments = list(set(s.assessment_type for s in subj_scores))
                subj_terms = list(set(s.term for s in subj_scores))
                students_scored = len(set(s.student_id for s in subj_scores))
                subject_progress[subj] = {
                    "assessments_done": subj_assessments,
                    "terms_done": subj_terms,
                    "students_scored": students_scored
                }
                
                # Figure out next logical assessment
                standard_order = ['1st CA', '2nd CA', 'Open Day', 'Note Book', 'Assignment', 'Exam']
                missing = [a for a in standard_order if a not in subj_assessments]
                if missing:
                    grading_progress.append("{} → {} needs {} next for {}".format(
                        c.name, subj, missing[0],
                        subj_terms[0] if subj_terms else "1st Term"
                    ))
            
            class_data[c.name] = {
                "student_count": len(students),
                "students": [s.name for s in students[:MAX_STUDENTS_IN_CONTEXT]],
                "assessments": assessment_types,
                "subjects": subjects,
                "terms": terms,
                "subject_progress": subject_progress,
                "has_scores": len(scores) > 0
            }
        
        # Build current session analytics if results are available
        session_analytics = ""
        if current_results:
            valid_scores = []
            missing_names = []
            for r in current_results:
                name = r.get('name', '').strip()
                score = r.get('score', '').strip()
                if not name:
                    missing_names.append(r)
                    continue
                try:
                    val = str(score)
                    if '/' in val: val = val.split('/')[0]
                    numeric = float(val)
                    valid_scores.append({"name": name, "score": numeric, "class": r.get('class', '')})
                except:
                    pass
            
            if valid_scores:
                scores_list = [s['score'] for s in valid_scores]
                avg = sum(scores_list) / len(scores_list)
                highest = max(valid_scores, key=lambda x: x['score'])
                lowest = min(valid_scores, key=lambda x: x['score'])
                
                # Anomaly detection — scores that deviate significantly from average
                anomalies = []
                if len(scores_list) > 3:
                    import statistics
                    stdev = statistics.stdev(scores_list) if len(scores_list) > 1 else 0
                    for s in valid_scores:
                        if stdev > 0 and abs(s['score'] - avg) > (2 * stdev):
                            anomalies.append(s)
                
                session_analytics = """
CURRENT GRADING SESSION ANALYTICS:
- Total students scanned: {total}
- Average score: {avg:.1f}
- Highest: {high_name} ({high_score})
- Lowest: {low_name} ({low_score})
- Students with missing names: {missing}
- Anomalies (unusual scores): {anomalies}
- Score distribution: {dist}
""".format(
                    total=len(valid_scores),
                    avg=avg,
                    high_name=highest['name'], high_score=highest['score'],
                    low_name=lowest['name'], low_score=lowest['score'],
                    missing=len(missing_names),
                    anomalies=json.dumps([{"name": a['name'], "score": a['score']} for a in anomalies]) if anomalies else "None detected",
                    dist="Below avg: {}, Above avg: {}".format(
                        len([s for s in scores_list if s < avg]),
                        len([s for s in scores_list if s >= avg])
                    )
                )
        
        system_prompt = """You are the Smart Assistant for QSI Smart Grader, designed by XO.
You are a highly capable AI teaching aide. You KNOW you can read handwriting, scan images, extract grades, merge them into Nigerian mark books, calculate totals, and identify anomalies.

PERSONALITY: Fluid, natural, concise, and genuinely helpful. Never sound robotic, overly enthusiastic, or rigid. Don't act clueless or ask unnecessary confirmation questions. If a user asks you to do something you know how to do (like scanning a script), just execute the action immediately instead of asking for confirmation. Talk like a brilliant, no-nonsense colleague.

NIGERIAN MARK BOOK RULES (you MUST know these):
- Columns 1-5 are CAs: 1st CA (10), 2nd CA (10), then 3 flexible columns from [Open Day, Note Book, Assignment, Attendance]
- 4 items fill 3 slots, so ONE column combines two items and is worth 20 (e.g., "Open Day/Attendance")
- 1st CA and 2nd CA can NEVER be the 20-point column
- Column 6 = Total CA = sum(columns 1-5) / 2 = max 30
- Column 7 = Exam = max 70
- Column 8 = Grand Total = Total CA + Exam = max 100
- 2nd Term: includes "1st Term Total" + "1st & 2nd" (sum of both Grand Totals)
- 3rd Term: includes "1st Term Total" + "2nd Term Total" + "1st 2nd & 3rd" (sum of all 3) + "Average" (sum / 3) + "Position"
- There is NO separate "Annual" sheet. The 3rd Term IS the annual summary.

CURRENT STATE:
- Teacher is on the "{screen}" screen
- Session: {session}
- Context: {context}

DATABASE (live rosters):
{db_data}

{analytics}

{progress}

ACTIONS YOU CAN TAKE (pick the best one):
{{
    "response": "Your warm, natural response (1-3 sentences). Use real names and data.",
    "action": One of:
        "setup_session" - Set up class/subject/assessment, navigate to scanning
        "view_standings" - Show results and rankings
        "add_student" - Add a student to a class roster
        "add_students_batch" - Add multiple students at once
        "move_student" - Move a student between classes
        "correct_score" - Fix a specific student's score in the current session
        "edit_scores" - Go to review screen to fix scores
        "add_class" - Open the add-class form
        "update_roster" - Open class list for name fixes
        "export_data" - Download the Excel file
        "analyze_scores" - Show analytics card (averages, trends, insights)
        "compare_classes" - Side-by-side class performance comparison
        "compare_assessments" - Compare across assessment types (1st CA vs 2nd CA)
        "flag_anomalies" - Highlight unusual scores that may be errors
        "find_at_risk" - List students who are failing or at risk
        "generate_report" - Create a formatted summary for admin/principal
        "edit_excel" - Edit an uploaded Excel file based on instructions (e.g. "add 5 marks to everyone")
        "scan_image_to_excel" - Extract requested columns from an uploaded image specifically into an Excel file.
        "none" - Just a conversational answer, no action needed
    "params": {{
        "class_name": "...",
        "subject_name": "...",
        "assessment_type": "...",
        "student_name": "...",
        "target_class": "...",
        "new_score": "...",
        "instruction": "...",
        "students": ["name1", "name2"],
        "report_text": "...",
        "insights": [list of insight strings],
        "anomalies": [list of {{name, score, reason}}],
        "at_risk": [list of {{name, score, class}}]
    }}
}}

CRITICAL ACTION SELECTION RULES:
>>> NEVER return action "none" when the teacher is asking you to DO something. <<<
>>> If the teacher wants to grade, scan, start a session, add a test → use "setup_session" <<<
>>> If the teacher mentions a student name + score → use "correct_score" immediately <<<
>>> If the teacher says "yes", "ok", "do it", "go ahead" → EXECUTE the action from the previous exchange, don't just talk about it <<<
>>> ACTION FIRST, CHAT SECOND. Always pick an action. Only use "none" for pure questions like "what is this app?" <<<

EXAMPLE INPUT→OUTPUT MAPPINGS (follow these patterns exactly):
- "I want to add 2nd test" → action: "setup_session", params: {{assessment_type: "2nd CA"}}
- "Grade next test" → action: "setup_session" with next logical assessment
- "Add 2nd CA for SS 1Q" → action: "setup_session", params: {{class_name: "SS 1Q", assessment_type: "2nd CA"}}
- "Adekunie should be 8" → action: "correct_score", params: {{student_name: "Adekunie", new_score: "8"}}
- "Fix Tunde's score to 15" → action: "correct_score", params: {{student_name: "Tunde", new_score: "15"}}
- "Add 5 points to everyone" / "Delete students below 40" → action: "edit_excel", params: {{instruction: "add 5 points to everyone"}}
- "Make the excel our standard format" / "Convert to standard format" / "Make it mark book format" → action: "edit_excel", params: {{instruction: "Restructure this Excel into the Nigerian school mark book standard format. The standard column order is: name, 1st CA, 2nd CA, Open Day, Note, Assignment, Total CA (sum of CAs ÷ 2, max 30), Exam (max 70), Grand Total (Total CA + Exam, max 100). Rename any detected score/assessment columns to match. Compute Total CA and Grand Total if raw CA and Exam scores exist. Keep the name column as-is."}}
- "Scan this image and extract Name, 1st CA, and 2nd CA into an excel file" → action: "scan_image_to_excel", params: {{instruction: "extract Name, 1st CA, and 2nd CA"}}
- "Scan this image for SS 1Q and extract Name, 1st CA, and 2nd CA" → action: "scan_image_to_excel", params: {{instruction: "extract Name, 1st CA, and 2nd CA", class_name: "SS 1Q"}}
- "This is SS 1T Mathematics, extract everything" → action: "scan_image_to_excel", params: {{instruction: "extract all columns", class_name: "SS 1T", subject_name: "Mathematics"}}
- "Extract all columns for 1st term" → action: "scan_image_to_excel", params: {{instruction: "extract all columns", assessment_type: "1st Term"}}
- "Just scan it" → action: "none" (ASK: "What columns should I extract? Or shall I grab everything I can see?")
- "Read this file" / "What is in this excel?" → action: "none" (Just read it and summarize conversationaly!)
- "Where is the edited file?" / "Did you edit it?" → action: "none" (Answer the question conversationally, DO NOT use edit_excel!)
- "Add Fatimah to SS 1Q" → action: "add_student", params: {{student_name: "Fatimah", class_name: "SS 1Q"}}
- "Only one student" or "For a student..." → action: "add_student"
- "Show results" / "See scores" → action: "view_standings"
- "Download" / "Get my Excel" → action: "export_data"
- "Who is failing?" → action: "find_at_risk"
- "Any issues with scores?" → action: "flag_anomalies"
- "Compare SS 1Q and SS 1S" → action: "compare_classes"
- "Set it up" / "yes" / "ok" / "do it" / "go ahead" → REPEAT the previous action with same params

INTELLIGENCE RULES:
1. For "setup_session": Figure out the NEXT logical assessment. If 1st CA exists, suggest 2nd CA. If both exist, suggest Exam. ALWAYS include class_name and assessment_type in params.
2. For "correct_score": Extract the student name and new score from the message. Match fuzzy names to the roster.
3. For "analyze_scores": Share meaningful insights using the analytics data above.
4. For "add_students_batch": When teacher lists multiple names, extract ALL of them.
5. For "move_student": Include source class, student name, and destination class in params.
6. BE PROACTIVE: Mention data insights naturally.
7. Always reference real student names and data. Never make up data.
8. RESPONSE LENGTH: Keep it SHORT (1-2 sentences) WHEN confirming an app action. HOWEVER, if the teacher asks a general question, wants to EXPLAIN A FILE, needs an email drafted, or asks for a lesson plan, act as a full LLM and write as much as needed! Format long responses nicely.
9. CONVERSATION MEMORY: Use history context. "yes"/"ok"/"do it" = execute previous action.
10. ALWAYS return valid JSON with "response", "action" and "params". For general LLM tasks, use action "none" and put your full, rich answer in "response".
11. READ VS EDIT EXCEL: If the teacher asks you to simply "read", "review", or "tell me what you found" about an uploaded Excel file, DO NOT use "edit_excel". Only use "edit_excel" if they explicitly ask you to MODIFY data (e.g. add, delete, rename). If they just want to read, use action "none" and summarize it in the response based on the "Context" I provided you about the upload.
12. AVOID FALSE EDITS: If the teacher asks a question LIKE "where is the edited file?", they are asking a question, NOT giving an instruction to edit an Excel file. Use action "none".
13. IMAGE SCAN INTELLIGENCE: When teacher uploads image(s), you MUST collect ALL of the following BEFORE triggering scan_image_to_excel.
    a) WHICH CLASS is this for? → params.class_name (e.g. "SS 1T", "SS 2Q") — REQUIRED
    b) WHAT SUBJECT? → params.subject_name (e.g. "Mathematics") — REQUIRED
    c) WHICH TERM? → params.assessment_type (e.g. "1st Term", "2nd Term") — REQUIRED
    d) WHAT COLUMNS to extract? → params.instruction (e.g. "extract all columns")
    
    You can often detect class, subject, and term from the image header itself. If you can SEE it in the image, tell the teacher what you see and ask for confirmation.
    
    >>> CRITICAL RULE <<< 
    If ANY of the 3 required fields (class, subject, term) are missing and not visible in the image, you MUST ask the teacher for the missing info. 
    While you are asking a question to collect this info, your action MUST BE "none". 
    DO NOT use action "scan_image_to_excel" until you have confirmed ALL required information.
    
14. SMART COLUMN GUESSING: Nigerian record sheets often have these columns: 1st CA, 2nd CA, Open Day, Note Book, Assignment, Attendance, Total, Exam, Grand Total. If teacher says "extract all columns" or "everything", use these standard names.
15. AUTO-DETECT FROM IMAGE: If you can see the image AND it clearly shows a class name or subject in the header, TELL the teacher what you see and ask them to confirm. Be specific: "I can see this says 'SS 1T - Mathematics' at the top. Is that right?"
16. MULTI-CLASS UPLOADS: If teacher uploads multiple images and says they're for DIFFERENT classes, ask which image is for which class. Do NOT assume all images are the same class.
17. STANDARD FORMAT: When teacher says "standard format", "our format", "mark book format", or "convert to standard" — they want the Nigerian school mark book structure:
    - Column order: name, 1st CA (10), 2nd CA (10), Open Day (10-20), Note (10), Assignment (10), Total CA (CAs÷2, max 30), Exam (70), Grand Total (100)
    - Use action "edit_excel" with a detailed instruction to restructure the columns. The instruction should tell the AI editor to rename score columns to standard names, compute Total CA = sum(CAs)/2, and Grand Total = Total CA + Exam.
    - If the uploaded Excel only has "name" and "score" (one column), ask the teacher: "Which assessment is this score for? (1st CA, 2nd CA, Exam, etc.)"
18. IMAGE AWARENESS: If images are attached, I can SEE them. Describe what I observe in the images when relevant. If the image shows a record sheet, describe the structure I see (number of columns, visible headers, etc.).
19. CONVERSATION FLOW: When I just asked a question and the teacher gives a short answer (like a class name, "yes", or a subject name), I MUST interpret it as an answer to my question, NOT as a new unrelated command. Context matters!
20. ONE QUESTION AT A TIME: Never ask more than one question. Teachers are busy — keep the conversation flowing with one question, wait for the answer, then proceed.
21. FOLLOW-UP INTELLIGENCE: When the conversation history shows I just opened a roster editor / asked "which class is this for?" / asked about a subject, the teacher's next short reply IS the answer. Process it accordingly.

{conversation}

Return ONLY raw JSON. No markdown wrapping."""

        # Build progress text BEFORE the outer format() call to avoid nested {} issues
        if grading_progress:
            progress_text = "GRADING PROGRESS (proactively mention relevant items):\n" + "\n".join(grading_progress[:10])
        else:
            progress_text = "No grading progress tracked yet."
        
        system_prompt = system_prompt.format(
            screen=current_screen,
            session=json.dumps(session_info),
            context=json.dumps(context),
            db_data=json.dumps(class_data, indent=2),
            analytics=session_analytics,
            progress=progress_text,
            conversation=conversation_context
        )

        

        # Build content parts: system prompt + any images + user message
        content_parts = [system_prompt]
        if images:
            for img in images[:5]:  # Max 5 images
                try:
                    img_data = img.get('data', '')
                    img_mime = img.get('mime_type', 'image/jpeg')
                    if img_data:
                        content_parts.append({
                            "mime_type": img_mime,
                            "data": base64.b64decode(img_data)
                        })
                except Exception as img_err:
                    logger.warning("Smart assistant image decode error: {}".format(img_err))
        content_parts.append(message)
        
        # Try primary model, fallback to lighter model
        raw_text = None
        for model_name in [AI_MODEL_PRIMARY, AI_MODEL_FALLBACK]:
            try:
                raw_text = _call_gemini(model_name, content_parts)
                if raw_text:
                    break
            except Exception as model_err:
                logger.warning("Chat model {} failed, trying fallback: {}".format(model_name, model_err))
                continue
        
        if not raw_text:
            raise Exception("All AI models failed to respond")
        
        # ─── Robust JSON extraction ───────────────────────────────
        # Strip markdown code fences and common AI quirks
        raw_text = re_mod.sub(r'^```(?:json)?\s*', '', raw_text)
        raw_text = re_mod.sub(r'\s*```$', '', raw_text)
        raw_text = raw_text.strip()
        
        # Clean common JSON quirks from AI output
        # 1. Trailing commas before closing braces/brackets
        raw_text = re_mod.sub(r',\s*([}\]])', r'\1', raw_text)
        # 2. Single quotes to double quotes (careful: only around keys/values)
        # 3. Unescaped newlines inside strings
        
        # Try direct JSON parse first
        result = None
        try:
            result = json.loads(raw_text)
        except json.JSONDecodeError:
            # Fallback 1: Find the outermost JSON object with response+action
            json_match = re_mod.search(r'\{[\s\S]*"response"[\s\S]*"action"[\s\S]*\}', raw_text)
            if json_match:
                try:
                    cleaned = re_mod.sub(r',\s*([}\]])', r'\1', json_match.group())
                    result = json.loads(cleaned)
                except json.JSONDecodeError:
                    pass
            
            # Fallback 2: Try to find ANY JSON object
            if result is None:
                json_match2 = re_mod.search(r'\{[^{}]*"response"[^{}]*\}', raw_text)
                if json_match2:
                    try:
                        result = json.loads(json_match2.group())
                    except json.JSONDecodeError:
                        pass
            
            # Ultimate fallback: treat the entire response as conversational text
            if result is None:
                logger.warning("JSON parse failed, using text fallback. Raw: {}".format(raw_text[:300]))
                result = {
                    "response": raw_text,
                    "action": "none",
                    "params": {}
                }
        
        return jsonify(result), 200
        
    except Exception as e:
        import traceback
        logger.error("Smart assistant error: {}".format(e))
        logger.error(traceback.format_exc())
        
        # Categorize the error for a smart, specific response
        error_msg = str(e).lower()
        if 'quota' in error_msg or '429' in error_msg or 'resource' in error_msg:
            return jsonify({
                "response": "We're hitting the AI rate limit right now. Give it about 60 seconds and try again — I'll be ready!",
                "action": "none",
                "params": {}
            }), 200
        elif 'safety' in error_msg or 'blocked' in error_msg:
            return jsonify({
                "response": "The AI flagged that as potentially unsafe content. Could you rephrase your question?",
                "action": "none",
                "params": {}
            }), 200
        elif 'invalid' in error_msg or 'not found' in error_msg:
            return jsonify({
                "response": "Something went wrong with the AI model. This is temporary — please try again in a moment.",
                "action": "none",
                "params": {}
            }), 200
            
        return jsonify({
            "response": "Sorry, I had a hiccup processing that ({}: {}). Could you say it differently?".format(type(e).__name__, str(e)),
            "action": "none",
            "params": {}
        }), 200


@app.route('/api/safe-add-student', methods=['POST'])
def safe_add_student():
    """Safely add a student to a class roster with validation guards."""
    try:
        data = request.json
        student_name = str(data.get('studentName', '')).strip().title()
        class_name = str(data.get('className', '')).strip()
        force = data.get('force', False)  # Skip fuzzy check if user confirmed
        
        if not student_name or not class_name:
            return jsonify({"error": "Student name and class are required."}), 400
        
        # Guard 1: Validate class exists
        c = ClassModel.query.filter(func.lower(ClassModel.name) == class_name.lower()).first()
        if not c:
            return jsonify({
                "error": "Class '{}' not found. Please create the class first.".format(class_name),
                "suggestion": "add_class"
            }), 404
        
        # Guard 2: Check for exact duplicate
        existing = StudentModel.query.filter_by(class_id=c.id, name=student_name).first()
        if existing:
            return jsonify({
                "error": "{} is already in {}.".format(student_name, class_name),
                "duplicate": True
            }), 409
        
        # Guard 3: Fuzzy duplicate check - catch near-matches (skip if forced)
        all_students = StudentModel.query.filter_by(class_id=c.id).all()
        all_names = [s.name for s in all_students]
        if all_names and not force:
            best_match = process.extractOne(student_name, all_names, scorer=fuzz.token_set_ratio)
            if best_match and best_match[1] >= 88:
                return jsonify({
                    "warning": "A similar name exists: '{}' ({}% match). Is this the same student?".format(
                        best_match[0], best_match[1]
                    ),
                    "similar_name": best_match[0],
                    "similarity": best_match[1],
                    "needs_confirmation": True
                }), 200
        
        # All guards passed - safe to add
        new_student = StudentModel(class_id=c.id, name=student_name)
        db.session.add(new_student)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "{} has been added to {}.".format(student_name, class_name),
            "student_id": new_student.id,
            "class_name": class_name,
            "total_students": len(all_students) + 1
        }), 201
        
    except Exception as e:
        print("Safe add student error: {}".format(e))
        db.session.rollback()
        return jsonify({"error": str(e)}), 500



if __name__ == '__main__':
    # Make sure templates folder exists
    os.makedirs('templates', exist_ok=True)
    # Run server — debug mode only when explicitly enabled
    is_debug = os.getenv('FLASK_DEBUG', 'false').lower() in ('true', '1', 'yes')
    app.run(debug=is_debug, host='0.0.0.0', port=5000)
